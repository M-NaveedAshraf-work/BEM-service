from BEMP import BEM
import numpy as np
import pandas as pd
from scipy.stats import norm, lognorm
from pyDOE import lhs
from SALib.analyze import morris
# from app import UQData
import json
import openpyxl
import csv

#TODO: Need to pass UQdata in thorugh function not by calling it from app. It will not update
class UQ_Object(BEM):

    def __init__(self, buildingName, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Essol_30, Essol_45, Essol_60,
                 Essol_90, originalFile):

        # inherits BEM class
        BEM.__init__(self, buildingName, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Essol_30, Essol_45, Essol_60,
                     Essol_90)

        self.graph = True

        # read the "UQ_Input" Excel file
        # file = openpyxl.load_workbook('./Input/UQ_Input.xlsx', data_only=True)
        # file_sheet = file['UQ_Setting']
        #
        # self.UQ_mode = file_sheet.cell(row=2, column=3).value
        # self.num_of_sample = file_sheet.cell(row=3, column=3).value
        # self.output_utilized = file_sheet.cell(row=4, column=3).value
        UQData = originalFile
        self.UQ_mode = UQData['UQInputs']['UQMode']
        self.num_of_sample = int(UQData['UQInputs']['totalSamples'])
        self.output_utilized = UQData['UQInputs']['energyOutput']

        i = 0;
        self.param_info = {"num_vars": 0, "names": []}
        # while 1:
        #     if file_sheet.cell(row=i, column=3).value == None:
        #         break
        #     else:
        #         self.param_info["num_vars"] += 1
        #         self.param_info["names"].append(file_sheet.cell(row=i, column=3).value)
        #         i += 1
        state = 'go'
        while state != 'end':
            if UQData['UQParams'] == []:
                break
            else:
                self.param_info["num_vars"] += 1
                self.param_info["names"].append(UQData['UQParams'][i]['param'])
                i += 1
                if i >= len(UQData['UQParams']):
                    state = 'end'

        # i = 28;  ####
        # while 1:
        #     if file_sheet.cell(row=i, column=2).value == None:
        #         break
        #     else:
        #         self.param_info["num_vars"] += 1
        #         self.param_info["names"].append(
        #             "".join([file_sheet.cell(row=i, column=2).value, '_', file_sheet.cell(row=i, column=3).value]))
        #         i += 1

        i = 0;  ####
        while 1:
            print(UQData['heatParams'])
            if UQData['heatParams'] == []:
                break
            else:
                self.param_info["num_vars"] += 1
                self.param_info["names"].append(
                    "".join([UQData['heatParams'][i]['param'], '_', UQData['heatParams'][i]['month']]))
                i += 1
                if i >= len(UQData['heatParams']):
                    break

        # if self.UQ_mode == "Sensitivity Analysis":
        #     self.confidence_level = file_sheet.cell(row=5, column=3).value
        #     self.num_levels = file_sheet.cell(row=6, column=3).value

        if self.UQ_mode == "Sensitivity Analysis":
            self.confidence_level = float(UQData['UQInputs']['confidenceLevel'])
            self.num_levels = int(UQData['UQInputs']['numLevels'])

        # close the "UQ_Input" excel file
        # file.save('./Input/UQ_Input.xlsx')

        self.lhd = lhs(self.param_info["num_vars"], samples=self.num_of_sample)

    def LHS(self):
        # quantify variables' distributions using Latin HyperCube Sampling method
        """
        Distribution of each parameter (reference: UQ_Repository)
        1. Heating COP:               Normal distribution (s.d. 0.01)
        2. Cooling COP:               Normal distribution (s.d. 0.01)
        3. AirLeakageLevel:           WILL BE ADDED AFTER THE INFILTRATION PART IS UPDATED
        4. Occupancy:                 Normal distribution (s.d. 0.01)
        5. Appliance:                 Normal distribution (s.d. 0.01)
        6. Lighting:                  Normal distribution (s.d. 0.01)
        7. Outdoor Air:               Normal distribution (s.d. 0.01)
        8. Roof Uvalue:               Normal distribution (s.d. 0.01)
        9. Roof Absorptivity:         Normal distribution (s.d. 0.01)
        10. Roof Emissivity:          Normal distribution (s.d. 0.01)
        11. Opaque Uvalue:            Normal distribution (s.d. 0.01)
        12. Opaque Absorptivity:      Normal distribution (s.d. 0.01)
        13. Opaque Emissivity:        Normal distribution (s.d. 0.01)
        14. Window Uvalue:            Normal distribution (s.d. 0.01)
        15. Window Emissivity:        Normal distribution (s.d. 0.01)
        16. Window SolarTrasmittance: Normal distribution (s.d. 0.01)
        """
        j = 0  # j: column
        for param in self.param_info["names"]:
            if param == "HeatingCOP":
                self.lhd[:, j] = norm(loc=self.left_numeric[5, 0], scale=0.01).ppf(self.lhd[:, j])
            elif param == "CoolingCOP":
                self.lhd[:, j] = norm(loc=self.left_numeric[6, 0], scale=0.01).ppf(self.lhd[:, j])

            elif param == "AirLeakageLevel":  #### SHOULD BE MODIFIED LATER
                self.lhd[:, j] = norm(loc=self.left_numeric[8, 0], scale=0.01).ppf(self.lhd[:, j])

            elif param == "Occupancy":
                self.lhd[:, j] = norm(loc=self.zone[1, 0], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Appliance":
                self.lhd[:, j] = norm(loc=self.zone[3, 0], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Lighting":
                self.lhd[:, j] = norm(loc=self.zone[4, 0], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Outdoor Air":
                self.lhd[:, j] = norm(loc=self.zone[5, 0], scale=0.01).ppf(self.lhd[:, j])

            elif param == "Roof UValue":
                self.lhd[:, j] = norm(loc=self.material[0, 0], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Roof Absorption coefficient":
                self.lhd[:, j] = norm(loc=self.material[0, 1], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Roof Emissivity":
                self.lhd[:, j] = norm(loc=self.material[0, 2], scale=0.01).ppf(self.lhd[:, j])

            elif param == "Opaque UValue":
                self.lhd[:, j] = norm(loc=self.material[2, 0], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Opaque Absorption coefficient":
                self.lhd[:, j] = norm(loc=self.material[2, 1], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Opaque Emissivity":
                self.lhd[:, j] = norm(loc=self.material[2, 2], scale=0.01).ppf(self.lhd[:, j])

            elif param == "Window UValue":
                self.lhd[:, j] = norm(loc=self.material[4, 0], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Window Emissivity":
                self.lhd[:, j] = norm(loc=self.material[4, 2], scale=0.01).ppf(self.lhd[:, j])
            elif param == "Window Solar transmittance":
                self.lhd[:, j] = norm(loc=self.material[4, 3], scale=0.01).ppf(self.lhd[:, j])

            elif param == "Retail Refri Capacity":
                self.lhd[:, j] = norm(loc=self.retail_refrig_cap, scale=0.01).ppf(self.lhd[:, j])

            elif param == "EV Charger Power":
                self.lhd[:, j] = norm(loc=self.eb_capacity, scale=0.01).ppf(self.lhd[:, j])

            elif param == "Garage Appliance":
                self.lhd[:, j] = norm(loc=self.garage_appliance, scale=0.01).ppf(self.lhd[:, j])
            elif param == "Garage Lighting":
                self.lhd[:, j] = norm(loc=self.garage_lighting, scale=0.01).ppf(self.lhd[:, j])

            elif param == "Lighting Dimmer Weekday":
                self.lhd[:, j] = norm(loc=self.dimmer_wd, scale=0.01).ppf(self.lhd[:, j])
            elif param == "Lighting Dimmer Weekend":
                self.lhd[:, j] = norm(loc=self.dimmer_we, scale=0.01).ppf(self.lhd[:, j])

            elif param[:17] == "Occupancy Monthly":  ####
                if param[-3:] == "Jan":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[0, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Feb":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[1, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Mar":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[2, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Apr":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[3, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "May":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[4, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Jun":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[5, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Jul":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[6, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Aug":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[7, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Sep":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[8, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Oct":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[9, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Nov":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[10, 0], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Dec":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[11, 0], scale=0.01).ppf(self.lhd[:, j])

            elif param[:17] == "Appliance Monthly":
                if param[-3:] == "Jan":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[0, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Feb":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[1, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Mar":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[2, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Apr":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[3, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "May":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[4, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Jun":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[5, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Jul":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[6, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Aug":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[7, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Sep":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[8, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Oct":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[9, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Nov":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[10, 1], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Dec":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[11, 1], scale=0.01).ppf(self.lhd[:, j])

            elif param[:16] == "Lighting Monthly":
                if param[-3:] == "Jan":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[0, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Feb":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[1, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Mar":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[2, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Apr":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[3, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "May":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[4, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Jun":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[5, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Jul":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[6, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Aug":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[7, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Sep":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[8, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Oct":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[9, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Nov":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[10, 2], scale=0.01).ppf(self.lhd[:, j])
                elif param[-3:] == "Dec":
                    self.lhd[:, j] = norm(loc=self.monthly_internal_heat_gain[11, 2], scale=0.01).ppf(self.lhd[:, j])
            j += 1

    def BEMP_Iterator(self):
        result_compilation = np.zeros((self.num_of_sample, 1))

        for i in range(self.num_of_sample):  # i: row, j: column
            # Open JSON instance
            # with open(self.jsonData) as f:
            #     data = json.load(f)
            # with open(self.buildingName) as f:
            #     data = json.load(f)
            data = self.jsonData

            for j in range(self.param_info["num_vars"]):
                # Change the parameter values in JSON instance
                if self.param_info["names"][j] == "HeatingCOP":
                    data["HeatingCOP"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "CoolingCOP":
                    data["CoolingCOP"] = self.lhd[i, j]

                elif self.param_info["names"][j] == "AirLeakageLevel":
                    data["AirLeakageLevel"] = self.lhd[i, j]

                elif self.param_info["names"][j] == "Occupancy":
                    data["Zone1"]["Occupancy"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Appliance":
                    data["Zone1"]["Appliance"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Lighting":
                    data["Zone1"]["Lighting"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "OutdoorAir":
                    data["Zone1"]["OutdoorAir"] = self.lhd[i, j]

                elif self.param_info["names"][j] == "Roof UValue":
                    data["Material"]["Roof1"]["UValue"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Roof Absorption coefficient":
                    data["Material"]["Roof1"]["Absorptivity"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Roof Emissivity":
                    data["Material"]["Roof1"]["Emissivity"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Opaque UValue":
                    data["Material"]["Opaque1"]["UValue"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Opaque Absorption coefficient":
                    data["Material"]["Opaque1"]["Absorptivity"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Opaque Emissivity":
                    data["Material"]["Opaque1"]["Emissivity"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Window UValue":
                    data["Material"]["Window1"]["UValue"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Window Emissivity":
                    data["Material"]["Window1"]["Emissivity"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Window Solar transmittance":
                    data["Material"]["Window1"]["SolarTrasmittance"] = self.lhd[i, j]


                elif self.param_info["names"][j] == "Retail Refri Capacity":  ####
                    data["RetailRefrig"]["Capacity"] = self.lhd[i, j]

                elif self.param_info["names"][j] == "EV Charger Power":
                    data["ElectricVehicle"]["ChargingPower"] = self.lhd[i, j]

                elif self.param_info["names"][j] == "Garage Appliance":
                    data["Garage"]["Appliance"] = self.lhd[i, j]

                elif self.param_info["names"][j] == "Garage Lighting":
                    data["Garage"]["Lighting"] = self.lhd[i, j]

                elif self.param_info["names"][j] == "Lighting Dimmer Weekday":
                    data["LightingDimmer"]["DimmerSaving_WD"] = self.lhd[i, j]
                elif self.param_info["names"][j] == "Lighting Dimmer Weekend":
                    data["LightingDimmer"]["DimmerSaving_WE"] = self.lhd[i, j]

                elif self.param_info["names"][j][:17] == "Occupancy Monthly":
                    if self.param_info["names"][j][-3:] == "Jan":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Jan"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Feb":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Feb"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Mar":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Mar"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Apr":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Apr"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "May":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["May"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Jun":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Jun"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Jul":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Jul"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Aug":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Aug"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Sep":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Sep"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Oct":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Oct"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Nov":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Nov"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Dec":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Dec"] = self.lhd[i, j]


                elif self.param_info["names"][j][:17] == "Appliance Monthly":
                    if self.param_info["names"][j][-3:] == "Jan":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Jan"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Feb":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Feb"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Mar":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Mar"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Apr":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Apr"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "May":
                        data["Monthly_InternalHeatGain"]["Appliance"]["May"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Jun":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Jun"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Jul":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Jul"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Aug":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Aug"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Sep":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Sep"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Oct":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Oct"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Nov":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Nov"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Dec":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Dec"] = self.lhd[i, j]

                elif self.param_info["names"][j][:16] == "Lighting Monthly":
                    if self.param_info["names"][j][-3:] == "Jan":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Jan"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Feb":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Feb"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Mar":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Mar"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Apr":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Apr"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "May":
                        data["Monthly_InternalHeatGain"]["Lighting"]["May"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Jun":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Jun"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Jul":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Jul"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Aug":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Aug"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Sep":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Sep"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Oct":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Oct"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Nov":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Nov"] = self.lhd[i, j]
                    elif self.param_info["names"][j][-3:] == "Dec":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Dec"] = self.lhd[i, j]

            # Save the JSON instance
            # with open(self.buildingName, 'w', encoding='utf-8') as f:
            #     json.dump(data, f, ensure_ascii=False, indent=4)
            # with open(self.jsonData, 'w', encoding='utf-8') as f:
            #     json.dump(data, f, ensure_ascii=False, indent=4)

            # Re-iterations
            self.readJSON()
            self.GeneralDataProcessing()
            self.TransmissionHeatTransfer()
            self.VentilationAirFlowandVentilationHeatTransfer()
            self.PumpSystemEnergy()
            self.DHWandSolarWaterHeating()
            outcome, outcome2, outcome3 = self.hourly_BEM()

            if self.output_utilized == "Delivered Energy":
                result_compilation[i, 0] = outcome2  # unit: kWh/m2
            elif self.output_utilized == "Electricity Use":
                result_compilation[i, 0] = np.sum(outcome3[:, 0])  # unit: kWh
            elif self.output_utilized == "Natural Gas Use":
                result_compilation[i, 0] = np.sum(outcome3[:, 1])  # unit: kWh

            if i % 50 == 0:  # print iteration status at every 50 iterations
                print(f"{i}th iteration is complete.")

        return result_compilation

    def SAUA(self):

        if self.UQ_mode == "Sensitivity Analysis":
            # conduct the Morris method
            self.Y = self.BEMP_Iterator()
            print(self.param_info)
            print(self.lhd)
            print(self.Y)
            print(self.Y.size)
            print(self.confidence_level)
            print(self.num_levels)
            SA_result = morris.analyze(self.param_info, self.lhd, self.Y, conf_level=self.confidence_level,
                                       print_to_console=False, num_levels = self.num_levels)
            print(SA_result)
            # save the result in CSV
            output = []
            for i in range(len(SA_result["names"])):
                output.append(
                    [SA_result["names"][i], SA_result["mu"][i], SA_result["mu_star"][i], SA_result["sigma"][i],
                     SA_result["mu_star_conf"][i]])

            with open("./Output/Sensitivity_Analysis_Result.csv", 'w', newline='') as csvfile:
                csvwriter = csv.writer(csvfile)
                csvwriter.writerow(["Parameter", "mu", "mu_star", "sigma", "mu_star_conf"])
                for i in range(len(SA_result["names"])):
                    csvwriter.writerow(output[i])

            # visualization
            if self.graph == True:
                SA_result = pd.DataFrame.from_dict(dict(SA_result))

                import matplotlib.pyplot as plt
                fig, (ax1, ax2) = plt.subplots(1, 2, )
                fig.set_figheight(5)
                fig.set_figwidth(13)

                new_data = SA_result.sort_values(["mu_star"], ascending=True)
                # ax1.barh(new_data["names"], new_data["mu_star"].values, height=0.6)
                # ax1.set_xlabel("mu*", fontsize=12)
                #
                # ax2.scatter(SA_result["mu_star"], SA_result["sigma"])
                # for i, label in enumerate(SA_result["names"]):
                #     ax2.annotate(label, (SA_result["mu_star"][i], SA_result["sigma"][i]))
                #
                # ax2.set_xlabel("mu*", fontsize=13)
                # ax2.set_ylabel("sigma", fontsize=13)
                # ax2.grid()
                # plt.show()
                firstGraphNames = new_data["names"]
                firstGraphData = new_data["mu_star"].values
                secondGraphNames = new_data["names"]
                secondGraphData = [i for i in zip(SA_result["mu_star"], SA_result["sigma"])]

        elif self.UQ_mode == "Uncertainty Analysis":
            UQ_result = self.BEMP_Iterator()

            # save the result in CSV
            df = pd.DataFrame(self.lhd, columns=self.param_info["names"])
            df = df.assign(output=UQ_result)  # add UQ_Result value
            df.to_csv("./Output/Uncertainty_Analysis_Result.csv")

            # visualization
            if self.graph == True:
                # import matplotlib.pyplot as plt
                # fig, (ax1, ax2) = plt.subplots(1, 2, )
                # fig.set_figheight(5)
                # fig.set_figwidth(13)
                #
                # ax1.hist(df["output"], color='b', bins=50)
                # ax1.set_xlabel(self.output_utilized, fontsize=13)
                # ax1.set_ylabel("Frequency", fontsize=13)
                # ax1.axvline(x = df.output.mean(), color='r', ymax =  .9, label = f'Mean = {np.round(df.output.mean(),2)}')
                #
                count, bins_count = np.histogram(df["output"])
                cdf = np.cumsum(count / sum(count))
                # ax2.plot(bins_count[1:], cdf, label="CDF", color='b', linewidth=2)
                # ax2.axvline(x = df.output.mean(), color='r', label = f'Mean = {np.round(df.output.mean(),2)}')
                # ax2.set_xlabel(self.output_utilized, fontsize=13)
                # ax2.set_ylabel("Cumulative Probability", fontsize=13)
                # ax2.grid()

                # handles, labels = ax1.get_legend_handles_labels()
                # fig.legend(handles, labels, loc='upper right')

                # plt.show()
                firstGraphNames = self.output_utilized
                firstGraphData = df["output"]
                A = []
                for i in range(len(firstGraphData)):
                    A.append([])
                    A[i].append(firstGraphData[i])
                firstGraphData = A
                secondGraphNames = cdf.tolist()
                secondGraphData = bins_count[1:]

        return firstGraphNames, firstGraphData, secondGraphNames, secondGraphData


def UQ(buildingName, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Esol_30, Esol_45, Esol_60, Esol_90, originalFile):
    instance = UQ_Object(buildingName, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Esol_30, Esol_45, Esol_60,
                         Esol_90, originalFile)
    instance.readJSON()
    instance.GeneralDataProcessing()
    instance.TransmissionHeatTransfer()
    instance.VentilationAirFlowandVentilationHeatTransfer()
    instance.BEMSystem()
    instance.CoolingandHeatingSystemEnergy()
    instance.FanEnergy()
    instance.PumpSystemEnergy()
    instance.DHWandSolarWaterHeating()
    outcome, outcome2, outcome3 = instance.hourly_BEM()
    instance.LHS()
    firstGraphNames, firstGraphData, secoundGraphNames, secondGraphData = instance.SAUA()

    return firstGraphNames, firstGraphData, secoundGraphNames, secondGraphData


if __name__ == '__main__':
    print("Please execute the 'main.py' script.")
