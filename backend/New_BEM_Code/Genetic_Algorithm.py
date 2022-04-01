# from  CPython.BEMP import BEM
from BEMP import BEM
import numpy as np
import openpyxl
from random import random, randint, uniform, seed
from timeit import default_timer as timer
from math import sqrt
import csv
import json
import time
import pandas as pd
import matplotlib.pyplot as plt



class BEMP_Calibration_CapX(BEM):

    def __init__(self, jsonData, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Essol_30, Essol_45, Essol_60,
                 Essol_90, original_file_name, result_file_name):
        buildingName = "centergy_BEM_2019.json"
        # with open(buildingName) as f:
        #     data = json.load(f)
        data = jsonData
        calSettings = original_file_name
        self.original_file_name = "centergy_BEM_2019"
        buildingName = "".join(["./Input/", self.original_file_name, "_intermediate", ".json"])
        with open(buildingName, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        # Inherits BEM class
        BEM.__init__(self, jsonData, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Essol_30, Essol_45, Essol_60,
                     Essol_90)
        self.genetic_algorithm_setting = {}
        self.calibration_setting = {};
        self.CapX_setting = {}
        self.calibration_param_info = {}
        self.result_file_name = result_file_name
        # buildingName = "centergy_BEM_2019.json"
        self.buildingName = buildingName
        with open(self.buildingName) as f:
            data = json.load(f)
        building_name = "".join([buildingName[:-5], "_intermediate" ,".json"])
        with open(building_name, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        # 1. Open the "Calibration" excel file
        # file = openpyxl.load_workbook('./Input/BEM_Optimization_Input_'+original_file_name+'.xlsx', data_only=True)
        # file = openpyxl.load_workbook('./Input/BEM_Optimization_Input_v2_centergy_BEM_2019.xlsx', data_only=True)
        # file_sheet = file['GeneticAlgorithm_Setting']
        file_sheet = calSettings
        # 2. Read the genetic algorithm/calibration settings

        # Starting Modification to Read Calibration Settings

        # if file_sheet.cell(row=2, column=3).value < 2:
        #     raise Exception("The number of population must be greater than 1. Please input the number greater than 1.")
        # else:
        #     self.genetic_algorithm_setting["num_of_population"] = file_sheet.cell(row=2, column=3).value

        if file_sheet["ga_settings"]["population_size"] < 2:
            raise Exception("The number of population must be greater than 1. Please input the number greater than 1.")
        else:
            self.genetic_algorithm_setting["num_of_population"] = file_sheet["ga_settings"]["population_size"]

        # self.genetic_algorithm_setting["crossover_rate"] = file_sheet.cell(row=3, column=3).value
        # self.genetic_algorithm_setting["mutation_rate"] = file_sheet.cell(row=4, column=3).value
        # self.genetic_algorithm_setting["random_seed"] = file_sheet.cell(row=5, column=3).value
        # self.genetic_algorithm_setting["Code_Execution_Time"] = file_sheet.cell(row=6, column=3).value
        # self.genetic_algorithm_setting["calibration_or_CapX"] = file_sheet.cell(row=7, column=3).value
        # self.genetic_algorithm_setting["top_percentage_selection"] = file_sheet.cell(row=8, column=3).value

        self.genetic_algorithm_setting["crossover_rate"] = file_sheet["ga_settings"]["crossover_rate"]
        self.genetic_algorithm_setting["mutation_rate"] = file_sheet["ga_settings"]["mutation_rate"]
        self.genetic_algorithm_setting["random_seed"] = file_sheet["ga_settings"]["random_seed"]
        self.genetic_algorithm_setting["Code_Execution_Time"] = file_sheet["ga_settings"]["max_time"]
        self.genetic_algorithm_setting["calibration_or_CapX"] = file_sheet["type"]
        self.genetic_algorithm_setting["top_percentage_selection"] = file_sheet["ga_settings"][
            "top_percentage"]

        print(f'genetic_algorithm_setting: {self.genetic_algorithm_setting}')

        if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
            # self.calibration_setting["Data_interval"] = file_sheet.cell(row=10, column=3).value
            # self.calibration_setting["Elec_data"] = file_sheet.cell(row=11, column=3).value
            # self.calibration_setting["NaturalGas_data"] = file_sheet.cell(row=12, column=3).value
            # self.calibration_setting["cvRMSE_criterion"] = file_sheet.cell(row=13, column=3).value
            # self.calibration_setting["Convergence"] = file_sheet.cell(row=14, column=3).value
            # self.calibration_setting["N_times_Convergence"] = file_sheet.cell(row=15, column=3).value

            self.calibration_setting["Data_interval"] = file_sheet["calibration_settings"]["data_interval"]
            self.calibration_setting["Elec_data"] = file_sheet["calibration_settings"]["electricity_data"]
            self.calibration_setting["NaturalGas_data"] = file_sheet["calibration_settings"]["natural_gas_data"]
            self.calibration_setting["cvRMSE_criterion"] = file_sheet["calibration_settings"]["cvrmse_criterion"]
            self.calibration_setting["Convergence"] = file_sheet["calibration_settings"]["convergence_diff"]
            self.calibration_setting["N_times_Convergence"] = file_sheet["calibration_settings"]["n_times_converge"]

            # 3. Read the calibration parameter name/min/max
            # i = 19;
            i = 0;
            self.num_of_parameters = 0;
            self.calibration_parameters = []
            # while 1:
            #     if file_sheet.cell(row=i, column=3).value == None and file_sheet.cell(row=i, column=4).value == None:
            #         break
            #     elif file_sheet.cell(row=i, column=3).value != None and file_sheet.cell(row=i, column=4).value != None:
            #         self.calibration_param_info[file_sheet.cell(row=i, column=2).value] = {}
            #         self.calibration_param_info[file_sheet.cell(row=i, column=2).value]["Type"] = file_sheet.cell(row=i,
            #                                                                                                       column=3).value
            #         self.calibration_param_info[file_sheet.cell(row=i, column=2).value]["Min"] = file_sheet.cell(row=i,
            #                                                                                                      column=4).value
            #         self.calibration_param_info[file_sheet.cell(row=i, column=2).value]["Max"] = file_sheet.cell(row=i,
            #                                                                                                      column=5).value
            #
            #         if file_sheet.cell(row=i, column=4).value >= file_sheet.cell(row=i, column=5).value:
            #             raise Exception(
            #                 f"Max value {file_sheet.cell(row=i, column=5).value} must be greater than min value {file_sheet.cell(row=i, column=4).value}.")
            #
            #         self.num_of_parameters += 1
            #         self.calibration_parameters.append(file_sheet.cell(row=i, column=2).value)
            #     else:
            #         raise Exception("Please check the 'Calibration Parameters' table.")
            #     i += 1

            while 1:
                if file_sheet["calibration_parameters"][i] == []:
                    break
                elif file_sheet["calibration_parameters"][i] != []:
                    self.calibration_param_info[file_sheet["calibration_parameters"][i]["name"]] = {}
                    self.calibration_param_info[file_sheet["calibration_parameters"][i]["name"]]["Type"] = file_sheet["calibration_parameters"][i]["data"]
                    self.calibration_param_info[file_sheet["calibration_parameters"][i]["name"]]["Min"] = file_sheet["calibration_parameters"][i]["min"]
                    self.calibration_param_info[file_sheet["calibration_parameters"][i]["name"]]["Max"] = file_sheet["calibration_parameters"][i]["max"]

                    if file_sheet["calibration_parameters"][i]["min"] >= file_sheet["calibration_parameters"][i]["max"]:
                        raise Exception(
                            f"Max value must be greater than min value.")

                    self.num_of_parameters += 1
                    self.calibration_parameters.append(file_sheet["calibration_parameters"][i]["name"])
                else:
                    raise Exception("Please check the 'Calibration Parameters' table.")
                i += 1

            # 3.1 Read the calibration SCHEDULE parameters
            import random
            self.randomlist = list(
                map(str, random.sample(range(10, 20), 6)))  # generate 6 2-digit integers in str format
            # i = 46;
            # k = 0
            # while 1:
            #     if file_sheet.cell(row=i, column=3).value == None and file_sheet.cell(row=i, column=4).value == None:
            #         break
            #     elif file_sheet.cell(row=i, column=3).value != None and file_sheet.cell(row=i, column=4).value != None:
            #         self.calibration_param_info[self.randomlist[k]] = {}
            #         self.calibration_param_info[self.randomlist[k]]["Name"] = file_sheet.cell(row=i, column=2).value
            #         self.calibration_param_info[self.randomlist[k]]["Type"] = file_sheet.cell(row=i, column=3).value
            #         self.calibration_param_info[self.randomlist[k]]["Hour_Min"] = file_sheet.cell(row=i, column=4).value
            #         self.calibration_param_info[self.randomlist[k]]["Hour_Max"] = file_sheet.cell(row=i, column=5).value
            #         self.calibration_param_info[self.randomlist[k]]["Min"] = file_sheet.cell(row=i, column=6).value
            #         self.calibration_param_info[self.randomlist[k]]["Max"] = file_sheet.cell(row=i, column=7).value
            #
            #         if file_sheet.cell(row=i, column=6).value >= file_sheet.cell(row=i, column=7).value:
            #             raise Exception(
            #                 f"Max value {file_sheet.cell(row=i, column=7).value} must be greater than min value {file_sheet.cell(row=i, column=6).value}.")
            #
            #         self.num_of_parameters += 1
            #         self.calibration_parameters.append(self.randomlist[k])
            #     else:
            #         raise Exception("Please check the 'Schedule Calibration Parameters' table.")
            #     i += 1;
            #     k += 1

            i = 0;
            k = 0
            while 1:
                if file_sheet["schedule_parameters"][i] == []:
                    break
                elif file_sheet["schedule_parameters"][i] != []:
                    self.calibration_param_info[self.randomlist[k]] = {}
                    self.calibration_param_info[self.randomlist[k]]["Name"] = file_sheet["schedule_parameters"][i]["name"]
                    self.calibration_param_info[self.randomlist[k]]["Type"] = file_sheet["schedule_parameters"][i]["day"]
                    self.calibration_param_info[self.randomlist[k]]["Hour_Min"] = file_sheet["schedule_parameters"][i]["min_hour"]
                    self.calibration_param_info[self.randomlist[k]]["Hour_Max"] = file_sheet["schedule_parameters"][i]["max_hour"]
                    self.calibration_param_info[self.randomlist[k]]["Min"] = file_sheet["schedule_parameters"][i]["min"]
                    self.calibration_param_info[self.randomlist[k]]["Max"] = file_sheet["schedule_parameters"][i]["max"]

                    if file_sheet["schedule_parameters"][i]["min"] >= file_sheet["schedule_parameters"][i]["max"]:
                        raise Exception(
                            f"Max value must be greater than min value.")

                    self.num_of_parameters += 1
                    self.calibration_parameters.append(self.randomlist[k])
                else:
                    raise Exception("Please check the 'Schedule Calibration Parameters' table.")
                i += 1;
                k += 1

            # 3.2 Read the calibration monthly internal heat gain parameters ####
            self.month_list = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            # i = 65
            # x = 1
            # y = 1
            # z = 1
            # while 1:
            #     if file_sheet.cell(row=i, column=3).value == None and file_sheet.cell(row=i, column=4).value == None:
            #         break
            #     elif file_sheet.cell(row=i, column=3).value != None and file_sheet.cell(row=i, column=4).value != None:
            #         if file_sheet.cell(row=i, column=2).value not in self.calibration_param_info:
            #             tempo_name = file_sheet.cell(row=i, column=2).value
            #             self.calibration_param_info[tempo_name] = {}
            #         else:  # if name already exists
            #             if file_sheet.cell(row=i, column=2).value[:9] == "Occupancy":
            #                 tempo_name = "".join([file_sheet.cell(row=i, column=2).value, "_", str(x)])
            #                 self.calibration_param_info[tempo_name] = {}
            #                 x += 1
            #
            #             elif file_sheet.cell(row=i, column=2).value[:9] == "Appliance":
            #                 tempo_name = "".join([file_sheet.cell(row=i, column=2).value, "_", str(y)])
            #                 self.calibration_param_info[tempo_name] = {}
            #                 y += 1
            #
            #             elif file_sheet.cell(row=i, column=2).value[:8] == "Lighting":
            #                 tempo_name = "".join([file_sheet.cell(row=i, column=2).value, "_", str(z)])
            #                 self.calibration_param_info[tempo_name] = {}
            #                 z += 1
            #
            #         self.calibration_param_info[tempo_name]["Type"] = file_sheet.cell(row=i, column=3).value
            #         self.calibration_param_info[tempo_name]["Min"] = file_sheet.cell(row=i, column=4).value
            #         self.calibration_param_info[tempo_name]["Max"] = file_sheet.cell(row=i, column=5).value
            #         self.calibration_param_info[tempo_name]["Month_Min"] = file_sheet.cell(row=i, column=6).value
            #         self.calibration_param_info[tempo_name]["Month_Max"] = file_sheet.cell(row=i, column=7).value
            #
            #         if file_sheet.cell(row=i, column=4).value >= file_sheet.cell(row=i, column=5).value:
            #             raise Exception(
            #                 f"Max value {file_sheet.cell(row=i, column=5).value} must be greater than min value {file_sheet.cell(row=i, column=4).value}.")
            #
            #         self.num_of_parameters += 1
            #         self.calibration_parameters.append(tempo_name)
            #     else:
            #         raise Exception("Please check the 'Monthly Internal Heat Gains' table.")
            #     i += 1

            i = 0
            x = 1
            y = 1
            z = 1
            while 1:
                if file_sheet["monthly_internal"][i] == []:
                    break
                elif file_sheet["monthly_internal"][i] != []:
                    if file_sheet["monthly_internal"][i]["name"] not in self.calibration_param_info:
                        tempo_name = file_sheet["monthly_internal"][i]["name"]
                        self.calibration_param_info[tempo_name] = {}
                    else:  # if name already exists
                        if file_sheet["monthly_internal"][i]["name"] == "Occupancy":
                            tempo_name = "".join([file_sheet["monthly_internal"][i]["name"], "_", str(x)])
                            self.calibration_param_info[tempo_name] = {}
                            x += 1

                        elif file_sheet["monthly_internal"][i]["name"] == "Appliance":
                            tempo_name = "".join([file_sheet["monthly_internal"][i]["name"], "_", str(y)])
                            self.calibration_param_info[tempo_name] = {}
                            y += 1

                        elif file_sheet["monthly_internal"][i]["name"] == "Lighting":
                            tempo_name = "".join([file_sheet["monthly_internal"][i]["name"], "_", str(z)])
                            self.calibration_param_info[tempo_name] = {}
                            z += 1

                    self.calibration_param_info[tempo_name]["Type"] = file_sheet["monthly_internal"][i]["day"]
                    self.calibration_param_info[tempo_name]["Min"] = file_sheet["monthly_internal"][i]["min"]
                    self.calibration_param_info[tempo_name]["Max"] = file_sheet["monthly_internal"][i]["max"]
                    self.calibration_param_info[tempo_name]["Month_Min"] = file_sheet["monthly_internal"][i]["min_month"]
                    self.calibration_param_info[tempo_name]["Month_Max"] = file_sheet["monthly_internal"][i]["max_month"]

                    if file_sheet["monthly_internal"][i]["min"] >= file_sheet["monthly_internal"][i]["max"]:
                        raise Exception(
                            f"Max value must be greater than min value.")

                    self.num_of_parameters += 1
                    self.calibration_parameters.append(tempo_name)
                else:
                    raise Exception("Please check the 'Monthly Internal Heat Gains' table.")
                i += 1

            # print(f"calibration_param_info:\n{self.calibration_param_info}")
            # print(f"calibration_parameters:\n{self.calibration_parameters}")

            # 4. Read the input elec/natural gas data
            if self.calibration_setting["Data_interval"] == "Monthly":
                self.measuredData = np.zeros((12, 2))
                self.data_sheet = file['Calibration_Monthly_Data']
                self.num_of_loop = 12  # needed in the "Evaluation" method

                for i in range(12):
                    if self.calibration_setting["Elec_data"] == "Yes":
                        self.measuredData[i, 0] = self.data_sheet.cell(row=i + 3, column=2).value
                    if self.calibration_setting["NaturalGas_data"] == "Yes":
                        self.measuredData[i, 1] = self.data_sheet.cell(row=i + 3, column=3).value

            elif self.calibration_setting["Data_interval"] == "Daily":
                self.measuredData = np.zeros((365, 2))
                self.data_sheet = file['Calibration_Daily_Data']
                self.num_of_loop = 365

                for i in range(365):
                    if self.calibration_setting["Elec_data"] == "Yes":
                        self.measuredData[i, 0] = self.data_sheet.cell(row=i + 3, column=2).value
                    if self.calibration_setting["NaturalGas_data"] == "Yes":
                        self.measuredData[i, 1] = self.data_sheet.cell(row=i + 3, column=3).value

            elif self.calibration_setting["Data_interval"] == "Hourly":
                self.measuredData = np.zeros((8760, 2))
                self.data_sheet = file['Calibration_Hourly_Data']
                self.num_of_loop = 8760

                for i in range(8760):
                    if self.calibration_setting["Elec_data"] == "Yes":
                        self.measuredData[i, 0] = self.data_sheet.cell(row=i + 3, column=3).value
                    if self.calibration_setting["NaturalGas_data"] == "Yes":
                        self.measuredData[i, 1] = self.data_sheet.cell(row=i + 3, column=4).value
            else:
                raise Exception("Please check the 'Calibration Data Interval' input.")

            # 5. Calculate the average data value
            self.y_bar = 0;
            self.number_of_data = 1  # y_bar: average utility data -> needed for cvRMSE calculation
            if self.calibration_setting["Elec_data"] == "Yes" and self.calibration_setting["NaturalGas_data"] == "Yes":
                self.y_bar = self.measuredData.mean()
                self.number_of_data = 2

            elif self.calibration_setting["Elec_data"] == "Yes" and self.calibration_setting["NaturalGas_data"] == "No":
                self.y_bar = self.measuredData.mean(axis=0)[0]

            elif self.calibration_setting["Elec_data"] == "No" and self.calibration_setting["NaturalGas_data"] == "Yes":
                self.y_bar = self.measuredData.mean(axis=0)[1]

            # 6. Calculating All Data Formats for graphing in BEM Dashboard

            # Monthly
            self.MonthlyData = np.zeros((12, 2))
            self.data_sheet = file['Calibration_Monthly_Data']
            self.num_of_loop_month = 12  # needed in the "Evaluation" method

            for i in range(12):
                if self.calibration_setting["Elec_data"] == "Yes":
                    self.MonthlyData[i, 0] = self.data_sheet.cell(row=i + 3, column=2).value
                if self.calibration_setting["NaturalGas_data"] == "Yes":
                    self.MonthlyData[i, 1] = self.data_sheet.cell(row=i + 3, column=3).value

            # Daily
            self.DailyData = np.zeros((365, 2))
            self.data_sheet = file['Calibration_Daily_Data']
            self.num_of_loop_day = 365

            for i in range(365):
                if self.calibration_setting["Elec_data"] == "Yes":
                    self.DailyData[i, 0] = self.data_sheet.cell(row=i + 3, column=2).value
                if self.calibration_setting["NaturalGas_data"] == "Yes":
                    self.DailyData[i, 1] = self.data_sheet.cell(row=i + 3, column=3).value

            # Hourly
            self.HourlyData = np.zeros((8760, 2))
            self.data_sheet = file['Calibration_Hourly_Data']
            self.num_of_loop_hour = 8760

            for i in range(8760):
                if self.calibration_setting["Elec_data"] == "Yes":
                    self.HourlyData[i, 0] = self.data_sheet.cell(row=i + 3, column=3).value
                if self.calibration_setting["NaturalGas_data"] == "Yes":
                    self.HourlyData[i, 1] = self.data_sheet.cell(row=i + 3, column=4).value


            #Calculate the average data value
            # self.y_bar = 0;
            self.number_of_data_real = 1  # y_bar: average utility data -> needed for cvRMSE calculation
            if self.calibration_setting["Elec_data"] == "Yes" and self.calibration_setting["NaturalGas_data"] == "Yes":
                self.y_bar_Monthly = self.MonthlyData.mean()
                self.y_bar_Daily = self.DailyData.mean()
                self.y_bar_Hourly = self.HourlyData.mean()
                self.number_of_data_real = 2

            elif self.calibration_setting["Elec_data"] == "Yes" and self.calibration_setting["NaturalGas_data"] == "No":
                self.y_bar_Monthly = self.MonthlyData.mean(axis=0)[0]
                self.y_bar_Daily = self.DailyData.mean(axis=0)[0]
                self.y_bar_Hourly = self.HourlyData.mean(axis=0)[0]

            elif self.calibration_setting["Elec_data"] == "No" and self.calibration_setting["NaturalGas_data"] == "Yes":
                self.y_bar_Monthly = self.MonthlyData.mean(axis=0)[1]
                self.y_bar_Daily = self.DailyData.mean(axis=0)[1]
                self.y_bar_Hourly = self.HourlyData.mean(axis=0)[1]


        elif self.genetic_algorithm_setting["calibration_or_CapX"] == "CapX":
            self.CapX_setting["Discount_Rate"]    = file_sheet.cell(row=10, column=10).value
            self.CapX_setting["Period"]           = file_sheet.cell(row=11, column=10).value
            self.CapX_setting["Budget"]           = file_sheet.cell(row=12, column=10).value
            self.CapX_setting["Delivered_Energy"] = file_sheet.cell(row=13, column=10).value
            self.CapX_setting["Elec_Cost"]        = file_sheet.cell(row=14, column=10).value
            self.CapX_setting["NaturalGas_Cost"]  = file_sheet.cell(row=15, column=10).value

            # 6. Read the CapX parameter name/min/max
            i = 19;
            self.CapX_input_parameters = {};
            self.num_of_parameters = 0;
            self.CapX_parameters = [];
            self.original_num_of_parameters = 0;
            self.CapX_parameters_no_duplicate = []
            self.discrete_param_set=["Heating and Cooling Plants efficiencies (COPs)", "Heat Recovery Type", "Roof1", "Opaque1", "Window1",\
                                     "Window1 Overhang Angle_S", "Window1 Overhang Angle_SE", "Window1 Overhang Angle_E", "Window1 Overhang Angle_NE", "Window1 Overhang Angle_N", "Window1 Overhang Angle_NW", "Window1 Overhang Angle_W", "Window1 Overhang Angle_SW",\
                                     "Window1 Fin Angle_S", "Window1 Fin Angle_SE", "Window1 Fin Angle_E", "Window1 Fin Angle_NE", "Window1 Fin Angle_N", "Window1 Fin Angle_NW", "Window1 Fin Angle_W", "Window1 Fin Angle_SW",\
                                     "Natural Ventilation", "Electric Battery", "Lighting Dimmer"] ####


            while 1:
                if file_sheet.cell(row=i, column=10).value == None:  ####
                    break
                elif file_sheet.cell(row=i, column=10).value != None:
                    self.CapX_input_parameters[file_sheet.cell(row=i, column=9).value] = file_sheet.cell(row=i,
                                                                                                         column=10).value

                    if file_sheet.cell(row=i, column=9).value in self.discrete_param_set and file_sheet.cell(row=i,
                                                                                                             column=10).value == "Continuous":
                        raise Exception(
                            "The type of 'Heating and Cooling Plants efficiencies (COPs)', 'Roof1', 'Opaque1', 'Window1', 'Overhang', 'Fin', 'SRF', 'Natural Ventilation', 'Electric Battery', 'Lighting Dimmer' must be discrete.")

                    if file_sheet.cell(row=i, column=9).value not in self.discrete_param_set:
                        self.num_of_parameters += 1
                        self.CapX_parameters.append(file_sheet.cell(row=i, column=9).value)

                    elif file_sheet.cell(row=i, column=9).value == "Heating and Cooling Plants efficiencies (COPs)":
                        self.num_of_parameters += 2
                        for j in range(2):
                            self.CapX_parameters.append(file_sheet.cell(row=i, column=9).value)
                    elif file_sheet.cell(row=i, column=9).value in ["Roof1", "Opaque1", "Window1",
                                                                    "Natural Ventilation", "Electric Battery"]:
                        self.num_of_parameters += 3
                        for j in range(3):
                            self.CapX_parameters.append(file_sheet.cell(row=i, column=9).value)
                    else:  # overhang, fin, lighting dimmer
                        self.num_of_parameters += 1
                        self.CapX_parameters.append(file_sheet.cell(row=i, column=9).value)  ####
                    self.original_num_of_parameters += 1
                else:
                    raise Exception("Please check the 'CapX Parameters' table.")
                i += 1
            self.CapX_parameters_no_duplicate = list(dict.fromkeys(self.CapX_parameters))

            # 7. Read parameter info
            self.CapX_param_table_info = {}
            row_number_pair = {"Heating COP": 36, "Cooling COP": 46, "Heating & Cooling COPs": 56,
                               "Heat Recovery Type": 66, "Building air leakage level": 76, \
                               "Specifiec fan power": 86, "DHW Generation System": 96, "PV module Area": 107,
                               "SHW Collector Area": 117, "Appliance": 127, "Lighting": 137, "Roof1": 147,
                               "Opaque1": 157, "Window1": 167, \
                               "Window1 Overhang Angle_S": 177, "Window1 Overhang Angle_SE": 185,
                               "Window1 Overhang Angle_E": 193, "Window1 Overhang Angle_NE": 201,
                               "Window1 Overhang Angle_N": 209, "Window1 Overhang Angle_NW": 217,
                               "Window1 Overhang Angle_W": 225, "Window1 Overhang Angle_SW": 233, \
                               "Window1 Fin Angle_S": 241, "Window1 Fin Angle_SE": 249, "Window1 Fin Angle_E": 257,
                               "Window1 Fin Angle_NE": 265, "Window1 Fin Angle_N": 273, "Window1 Fin Angle_NW": 281,
                               "Window1 Fin Angle_W": 289, "Window1 Fin Angle_SW": 297, "Wind Turbine": 305, \
                               "Natural Ventilation": 315, "Electric Battery": 325, "Lighting Dimmer": 335, \
                               "Window1 SRF_S": 343, "Window1 SRF_SE": 351, "Window1 SRF_E": 359, "Window1 SRF_NE": 367,
                               "Window1 SRF_N": 375, "Window1 SRF_NW": 383, "Window1 SRF_W": 391, \
                               "Window1 SRF_SW": 399, "Window1 SRF_Roof": 407}
            for param_name in self.CapX_input_parameters.keys():
                self.CapX_param_table_info[param_name] = []
                i = 0
                while 1:
                    tempo = {}
                    if file_sheet.cell(row=i + row_number_pair[param_name], column=10).value == None:  ####
                        break
                    else:
                        tempo["Name"] = file_sheet.cell(row=i + row_number_pair[param_name], column=9).value
                        tempo["Cost"] = file_sheet.cell(row=i + row_number_pair[param_name], column=10).value
                        tempo["Param1"] = file_sheet.cell(row=i + row_number_pair[param_name], column=11).value

                        if param_name == "Heating and Cooling Plants efficiencies (COPs)":
                            tempo["Param2"] = file_sheet.cell(row=i + row_number_pair[param_name], column=12).value
                        elif param_name in ["Roof1", "Opaque1", "Window1", "Natural Ventilation", "Electric Battery"]:
                            tempo["Param2"] = file_sheet.cell(row=i + row_number_pair[param_name], column=12).value
                            tempo["Param3"] = file_sheet.cell(row=i + row_number_pair[param_name],
                                                              column=13).value  ####

                        self.CapX_param_table_info[param_name].append(tempo)
                    i += 1

            self.interest_rate = (((1 + self.CapX_setting["Discount_Rate"]) ** self.CapX_setting["Period"]) - 1) / (
                        self.CapX_setting["Discount_Rate"] * (1 + self.CapX_setting["Discount_Rate"]) **
                        self.CapX_setting["Period"])
             # Equation: {((1+i)^n) -1}/(i*(1+i)^n), Ref: Engineering Economy, Blank and Tarquin, 7th ed. p.43

        ##            print(f"interest_rate: {self.interest_rate}")
        ##            print(f"CapX_input_parameters:\n{self.CapX_input_parameters}")
        ##            print(f"\nCapX_param_table_info:\n{self.CapX_param_table_info}")
        ##            print(f"\nCapX_parameters:\n{self.CapX_parameters}")
        ##            print(f"\nnum_of_parameters: {self.num_of_parameters}")
        ##            print(f"\noriginal_num_of_parameters: {self.original_num_of_parameters}")
        ##            print(f"\nCapX_parameters_no_duplicate: {self.CapX_parameters_no_duplicate}\n")

        # Close the "Calibration_Input" excel file
        # file.save('./Input/BEM_Optimization_Input.xlsx')

    # def JSON_Modification(self, chromosome, row_of_chromesome,building_name, calibrated = False):
    def JSON_Modification(self, chromosome, row_of_chromesome, calibrated=False):

        # Open JSON instnace
        # self.buildingName = "centergy_BEM_2019.json"
        with open(self.buildingName) as f:
            data = json.load(f)
        data = self.jsonData

        # Change the parameter values in JSON
        if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
            for j, name in enumerate(self.calibration_parameters):
                if name == "Heating COP":
                    data["HeatingCOP"] = chromosome[row_of_chromesome, j]
                elif name == "Cooling COP":
                    data["CoolingCOP"] = chromosome[row_of_chromesome, j]
                elif name == "Extra ventilation above fresh air supply":
                    data["ExtraVentilation"] = chromosome[row_of_chromesome, j]
                elif name == "Heat Recovery Type":
                    data["HeatRecoveryType"] = chromosome[row_of_chromesome, j]
                elif name == "Exhaust Air Recirculation Percentage":
                    data["ExhaustAirRecirculation"] = chromosome[row_of_chromesome, j]
                elif name == "Building air leakage level":
                    data["AirLeakageLevel"] = chromosome[row_of_chromesome, j]
                elif name == "Specific Fan Power":
                    data["SpecificFanPower"] = chromosome[row_of_chromesome, j]
                elif name == "Fan Control Factor":
                    data["FanControlFactor"] = chromosome[row_of_chromesome, j]
                elif name == "DHW Generation System":
                    data["DHWGenerationSystem"] = chromosome[row_of_chromesome, j]
                elif name == "Occupancy":
                    data["Zone1"]["Occupancy"] = chromosome[row_of_chromesome, j]
                elif name == "Appliance":
                    data["Zone1"]["Appliance"] = chromosome[row_of_chromesome, j]
                elif name == "Lighting":
                    data["Zone1"]["Lighting"] = chromosome[row_of_chromesome, j]
                elif name == "Outdoor Air":
                    data["Zone1"]["OutdoorAir"] = chromosome[row_of_chromesome, j]
                elif name == "DHW":
                    data["Zone1"]["DHW"] = chromosome[row_of_chromesome, j]

                elif name == "Roof1 Uvalue":
                    data["Material"]["Roof1"]["UValue"] = chromosome[row_of_chromesome, j]
                elif name == "Roof1 Absorption coefficient":
                    data["Material"]["Roof1"]["Absorptivity"] = chromosome[row_of_chromesome, j]
                elif name == "Roof1 Emissivity":
                    data["Material"]["Roof1"]["Emissivity"] = chromosome[row_of_chromesome, j]

                elif name == "Opaque1 Uvalue":
                    data["Material"]["Opaque1"]["UValue"] = chromosome[row_of_chromesome, j]
                elif name == "Opaque1 Absorption coefficient":
                    data["Material"]["Opaque1"]["Absorptivity"] = chromosome[row_of_chromesome, j]
                elif name == "Opaque1 Emissivity":
                    data["Material"]["Opaque1"]["Emissivity"] = chromosome[row_of_chromesome, j]

                elif name == "Window1 Uvalue":
                    data["Material"]["Window1"]["UValue"] = chromosome[row_of_chromesome, j]
                elif name == "Window1 Emissivity":
                    data["Material"]["Window1"]["Emissivity"] = chromosome[row_of_chromesome, j]
                elif name == "Window1 Solar transmittance":
                    data["Material"]["Window1"]["SolarTrasmittance"] = chromosome[row_of_chromesome, j]

                elif name[:11] == "Window1 SRF":
                    if name == "Window1 SRF All":
                        data["Envelope"]["Window1"]["S"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["SE"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["E"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["NE"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["N"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["NW"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["W"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["SW"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_S":
                        data["Envelope"]["Window1"]["S"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_SE":
                        data["Envelope"]["Window1"]["SE"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_E":
                        data["Envelope"]["Window1"]["E"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_NE":
                        data["Envelope"]["Window1"]["NE"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_N":
                        data["Envelope"]["Window1"]["N"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_NW":
                        data["Envelope"]["Window1"]["NW"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_W":
                        data["Envelope"]["Window1"]["W"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif name == "Window1 SRF_SW":
                        data["Envelope"]["Window1"]["SW"]["SRF"] = chromosome[row_of_chromesome, j]

                elif name == "Retail Refri Capacity":
                    data["RetailRefrig"]["Capacity"] = chromosome[row_of_chromesome, j]

                elif name == "EV Charger Power":
                    data["ElectricVehicle"]["ChargingPower"] = chromosome[row_of_chromesome, j]

                elif name == "Garage Appliance":
                    data["Garage"]["Appliance"] = chromosome[row_of_chromesome, j]

                elif name == "Garage Lighting":
                    data["Garage"]["Lighting"] = chromosome[row_of_chromesome, j]

                elif name == "Lighting Dimmer Weekday":
                    data["LightingDimmer"]["DimmerSaving_WD"] = chromosome[row_of_chromesome, j]

                elif name == "Lighting Dimmer Weekend":
                    data["LightingDimmer"]["DimmerSaving_WE"] = chromosome[row_of_chromesome, j]

                elif "Monthly" in name:  ####
                    if name[:9] == "Occupancy":
                        monthly_type = "Occupancy"
                    elif name[:9] == "Appliance":
                        monthly_type = "Appliance"
                    elif name[:8] == "Lighting":
                        monthly_type = "Lighting"

                    if self.month_list.index(self.calibration_param_info[name]["Month_Min"]) == self.month_list.index(
                            self.calibration_param_info[name]["Month_Max"]):
                        month = self.calibration_param_info[name]["Month_Min"]
                        data["Monthly_InternalHeatGain"][monthly_type][month] = chromosome[row_of_chromesome, j]

                    elif self.month_list.index(self.calibration_param_info[name]["Month_Min"]) < self.month_list.index(
                            self.calibration_param_info[name]["Month_Max"]):
                        # print(f"calibration_param_info[name]: {self.calibration_param_info[name]}")
                        for z in range(self.month_list.index(self.calibration_param_info[name]["Month_Min"]),
                                       self.month_list.index(self.calibration_param_info[name]["Month_Max"]) + 1):
                            data["Monthly_InternalHeatGain"][monthly_type][self.month_list[z]] = chromosome[
                                row_of_chromesome, j]
                    else:
                        raise Exception(
                            "Please check the 'Monthly_Min' and 'Monthly_Max' in the 'Monthly Internal Heat Gains' table.")
                #TODO: this was commented out below
                elif name[:9] == "Occupancy":
                    if name == "Occupancy Jan":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Jan"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Feb":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Feb"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Mar":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Mar"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Apr":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Apr"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy May":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["May"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Jun":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Jun"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Jul":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Jul"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Aug":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Aug"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Sep":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Sep"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Oct":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Oct"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Nov":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Nov"] = chromosome[row_of_chromesome, j]
                    elif name == "Occupancy Dec":
                        data["Monthly_InternalHeatGain"]["Occupancy"]["Dec"] = chromosome[row_of_chromesome, j]

                elif name[:9] == "Appliance":
                    if name == "Appliance Jan":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Jan"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Feb":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Feb"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Mar":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Mar"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Apr":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Apr"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance May":
                        data["Monthly_InternalHeatGain"]["Appliance"]["May"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Jun":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Jun"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Jul":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Jul"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Aug":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Aug"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Sep":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Sep"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Oct":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Oct"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Nov":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Nov"] = chromosome[row_of_chromesome, j]
                    elif name == "Appliance Dec":
                        data["Monthly_InternalHeatGain"]["Appliance"]["Dec"] = chromosome[row_of_chromesome, j]

                elif name[:8] == "Lighting":
                    if name == "Lighting Jan":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Jan"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Feb":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Feb"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Mar":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Mar"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Apr":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Apr"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting May":
                        data["Monthly_InternalHeatGain"]["Lighting"]["May"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Jun":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Jun"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Jul":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Jul"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Aug":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Aug"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Sep":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Sep"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Oct":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Oct"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Nov":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Nov"] = chromosome[row_of_chromesome, j]
                    elif name == "Lighting Dec":
                        data["Monthly_InternalHeatGain"]["Lighting"]["Dec"] = chromosome[row_of_chromesome, j]

                elif isinstance(int(name), int) == True:  # Temperature setpoint or Schedule
                    hour_max = self.calibration_param_info[name]["Hour_Max"]
                    hour_min = self.calibration_param_info[name]["Hour_Min"]
                    if self.calibration_param_info[name]["Hour_Max"] < self.calibration_param_info[name][
                        "Hour_Min"]:
                        hour_list = list(range(hour_min, 25)) + list(range(1, hour_max + 1))
                        hour_list.sort()
                    else:
                        hour_list = list(range(hour_min, hour_max + 1))
                    if self.calibration_param_info[name]["Name"] in ["Heating Temperature Setpoint",
                                                                     "Cooling Temperature Setpoint"]:
                        for k in hour_list:
                            data["TemperatureSetPoint"]["".join(["Hour",str(k)])]["".join([self.calibration_param_info[name]["Name"][:7], "_", self.calibration_param_info[name]["Type"]])] = chromosome[
                                row_of_chromesome, j]  ## added '_'

                    else:  # Occupancy, appliance, lighting

                        for k in hour_list:
                            data["Zone1_Schedule"]["".join(["Hour", str(k)])]["".join(
                                [self.calibration_param_info[name]["Name"][:-9], "_",
                                 self.calibration_param_info[name]["Type"]])] = chromosome[row_of_chromesome, j]



        else:  # CapX
            j = 0
            for param in self.CapX_parameters_no_duplicate:
                if param == "Heating COP":
                    data["HeatingCOP"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "Cooling COP":
                    data["CoolingCOP"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "Heating & Cooling COPs":
                    data["HeatingCOP"] = chromosome[row_of_chromesome, j]
                    data["CoolingCOP"] = chromosome[row_of_chromesome, j + 1]
                    j += 2
                elif param == "Heat Recovery Type":
                    data["HeatRecoveryType"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "Building air leakage level":
                    data["AirLeakageLevel"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "Specifiec fan power":
                    data["SpecificFanPower"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "DHW Generation System":
                    data["DHWGenerationSystem"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "PV module Area":
                    data["PVArea"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "SHW Collector Area":
                    data["SHWArea"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "Appliance":
                    data["Zone1"]["Appliance"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "Lighting":
                    data["Zone1"]["Lighting"] = chromosome[row_of_chromesome, j]
                    j += 1
                elif param == "Roof1":
                    data["Material"]["Roof1"]["UValue"] = chromosome[row_of_chromesome, j]
                    data["Material"]["Roof1"]["Absorptivity"] = chromosome[row_of_chromesome, j + 1]
                    data["Material"]["Roof1"]["Emissivity"] = chromosome[row_of_chromesome, j + 2]
                    j += 3
                elif param == "Opaque1":
                    data["Material"]["Opaque1"]["UValue"] = chromosome[row_of_chromesome, j]
                    data["Material"]["Opaque1"]["Absorptivity"] = chromosome[row_of_chromesome, j + 1]
                    data["Material"]["Opaque1"]["Emissivity"] = chromosome[row_of_chromesome, j + 2]
                    j += 3
                elif param == "Window1":
                    data["Material"]["Window1"]["UValue"] = chromosome[row_of_chromesome, j]
                    data["Material"]["Window1"]["Absorptivity"] = chromosome[row_of_chromesome, j + 1]
                    data["Material"]["Window1"]["SolarTrasmittance"] = chromosome[row_of_chromesome, j + 2]
                    j += 3
                elif param == "Natural Ventilation":
                    try:
                        data["NaturalVentilation"]["Width"]
                    except:
                        data["NaturalVentilation"] = {}

                    data["NaturalVentilation"]["Width"] = chromosome[row_of_chromesome, j]
                    data["NaturalVentilation"]["Height"] = chromosome[row_of_chromesome, j + 1]
                    data["NaturalVentilation"]["EffectiveAngle"] = chromosome[row_of_chromesome, j + 2]
                    j += 3

                elif param == "Electric Battery":
                    try:
                        data["ElectricBattery"]
                    except:
                        data["ElectricBattery"] = {}

                    data["ElectricBattery"]["Capacity"] = chromosome[row_of_chromesome, j]
                    data["ElectricBattery"]["ChargingEfficiency"] = chromosome[row_of_chromesome, j + 1]
                    data["ElectricBattery"]["DischargingEfficiency"] = chromosome[row_of_chromesome, j + 2]
                    data["ElectricBattery"]["ChargingPowerLimit"] = chromosome[row_of_chromesome, j] / 2
                    data["ElectricBattery"]["DischargingPowerLimit"] = chromosome[row_of_chromesome, j] / 2
                    j += 3
                elif param == "Lighting Dimmer":
                    tempo = 1 - (0.7 + 0.3 * (self.CapX_param_table_info["Lighting Dimmer"][-1]["Param1"] - chromosome[
                        row_of_chromesome, j]) / self.CapX_param_table_info["Lighting Dimmer"][-1]["Param1"])
                    data["LightingDimmer"]["DimmerSaving_WD"] = tempo  # Follows Yuna's thesis
                    data["LightingDimmer"]["DimmerSaving_WE"] = tempo
                    j += 1

                elif param[:16] == "Window1 Overhang Angle":
                    if param == "Window1 Overhang Angle_S":
                        data["Envelope"]["Window1"]["S"]["Overhang"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Overhang Angle_SE":
                        data["Envelope"]["Window1"]["SE"]["Overhang"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Overhang Angle_E":
                        data["Envelope"]["Window1"]["E"]["Overhang"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Overhang Angle_NE":
                        data["Envelope"]["Window1"]["NE"]["Overhang"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Overhang Angle_N":
                        data["Envelope"]["Window1"]["N"]["Overhang"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Overhang Angle_NW":
                        data["Envelope"]["Window1"]["NW"]["Overhang"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Overhang Angle_W":
                        data["Envelope"]["Window1"]["W"]["Overhang"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Overhang Angle_SW":
                        data["Envelope"]["Window1"]["SW"]["Overhang"] = chromosome[row_of_chromesome, j]
                    j += 1

                elif param[:17] == "Window1 Fin Angle":
                    if param == "Window1 Fin Angle_S":
                        data["Envelope"]["Window1"]["S"]["Fin"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Fin Angle_SE":
                        data["Envelope"]["Window1"]["SE"]["Fin"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Fin Angle_E":
                        data["Envelope"]["Window1"]["E"]["Fin"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Fin Angle_NE":
                        data["Envelope"]["Window1"]["NE"]["Fin"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Fin Angle_N":
                        data["Envelope"]["Window1"]["N"]["Fin"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Fin Angle_NW":
                        data["Envelope"]["Window1"]["NW"]["Fin"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Fin Angle_W":
                        data["Envelope"]["Window1"]["W"]["Fin"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 Fin Angle_SW":
                        data["Envelope"]["Window1"]["SW"]["Fin"] = chromosome[row_of_chromesome, j]
                    j += 1

                elif param[:17] == "Window1 SRF Angle":
                    if param == "Window1 SRF All":
                        data["Envelope"]["Window1"]["S"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["SE"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["E"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["NE"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["N"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["NW"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["W"]["SRF"] = chromosome[row_of_chromesome, j]
                        data["Envelope"]["Window1"]["SW"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_S":
                        data["Envelope"]["Window1"]["S"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_SE":
                        data["Envelope"]["Window1"]["SE"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_E":
                        data["Envelope"]["Window1"]["E"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_NE":
                        data["Envelope"]["Window1"]["NE"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_N":
                        data["Envelope"]["Window1"]["N"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_NW":
                        data["Envelope"]["Window1"]["NW"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_W":
                        data["Envelope"]["Window1"]["W"]["SRF"] = chromosome[row_of_chromesome, j]
                    elif param == "Window1 SRF Angle_SW":
                        data["Envelope"]["Window1"]["SW"]["SRF"] = chromosome[row_of_chromesome, j]
                    j += 1

        # Save the JSON instance
        # building_name = self.buildingName
        if calibrated:
            if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
                end_name = "_calibrated"
            else:
                end_name = "_capX"
            building_name = "".join(["./Input/", self.original_file_name, end_name, ".json"])
            with open(building_name, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        with open(self.buildingName, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    def Linear_Interpolation(self, x1, x2, y1, y2, x):
        return ((x - x1) * (y2 - y1) / (x2 - x1)) + y1

    def Calibration_Iteration(self, population):
        cvRMSE_compilation = []
        # building_name = "".join([self.buildingName[:-5], "_intermediate", ".json"])
        for i in range(self.genetic_algorithm_setting["num_of_population"]):
            # self.JSON_Modification(population, i, building_name = building_name)
            self.JSON_Modification(population, i)

            # 2. Re-calculations (Assumed BEM calculation was conducted at least once before the "cvRMSE" was executed.)
            self.readJSON()
            self.GeneralDataProcessing()
            self.TransmissionHeatTransfer()
            self.VentilationAirFlowandVentilationHeatTransfer()
            self.PumpSystemEnergy()
            self.DHWandSolarWaterHeating()

            outcome, outcome2, outcome3, grouped = self.hourly_BEM()  # outcome: self.deliveredEnergy, outcome2: self.Overall_deliveredEnergy, outcome3: self.deliveredEnergy_fuel
            """
            The fitness function is CVRMSE (Coefficient of variation of the Root Mean Square Error)
            CVRMSE = (1/y_bar) * sqrt(((sigma (yi - yhat_i)^2)/(n-p))
            where, yi: utility data
                   yi_bar: average utility data
                   yi_hat: simulation predicted data
                   n: the number of data points
                   p: 1 (usually)
            """

            if self.calibration_setting["Data_interval"] == "Monthly":
                if self.number_of_data == 2:
                    manipulated_result = np.zeros((12, 2))
                else:
                    manipulated_result = np.zeros((12, 1))

                if self.calibration_setting["Elec_data"] == "Yes":
                    manipulated_result[0, 0] = np.sum(outcome3[0:744, 0])
                    manipulated_result[1, 0] = np.sum(outcome3[744:1416, 0])
                    manipulated_result[2, 0] = np.sum(outcome3[1416:2159, 0])
                    manipulated_result[3, 0] = np.sum(outcome3[2159:2880, 0])
                    manipulated_result[4, 0] = np.sum(outcome3[2880:3624, 0])
                    manipulated_result[5, 0] = np.sum(outcome3[3624:4344, 0])
                    manipulated_result[6, 0] = np.sum(outcome3[4344:5088, 0])
                    manipulated_result[7, 0] = np.sum(outcome3[5088:5832, 0])
                    manipulated_result[8, 0] = np.sum(outcome3[5832:6552, 0])
                    manipulated_result[9, 0] = np.sum(outcome3[6552:7296, 0])
                    manipulated_result[10, 0] = np.sum(outcome3[7296:8016, 0])
                    manipulated_result[11, 0] = np.sum(outcome3[8016:8760, 0])

                if self.calibration_setting["NaturalGas_data"] == "Yes":
                    manipulated_result[0, 1] = np.sum(outcome3[0:744, 1])
                    manipulated_result[1, 1] = np.sum(outcome3[744:1416, 1])
                    manipulated_result[2, 1] = np.sum(outcome3[1416:2159, 1])
                    manipulated_result[3, 1] = np.sum(outcome3[2159:2880, 1])
                    manipulated_result[4, 1] = np.sum(outcome3[2880:3624, 1])
                    manipulated_result[5, 1] = np.sum(outcome3[3624:4344, 1])
                    manipulated_result[6, 1] = np.sum(outcome3[4344:5088, 1])
                    manipulated_result[7, 1] = np.sum(outcome3[5088:5832, 1])
                    manipulated_result[8, 1] = np.sum(outcome3[5832:6552, 1])
                    manipulated_result[9, 1] = np.sum(outcome3[6552:7296, 1])
                    manipulated_result[10, 1] = np.sum(outcome3[7296:8016, 1])
                    manipulated_result[11, 1] = np.sum(outcome3[8016:8760, 1])

            elif self.calibration_setting["Data_interval"] == "Daily":
                if self.number_of_data == 2:
                    manipulated_result = np.zeros((365, 2))
                else:
                    manipulated_result = np.zeros((365, 1))

                k = 0
                for m in range(0, 8760, 24):
                    if self.calibration_setting["Elec_data"] == "Yes":
                        manipulated_result[k, 0] = np.sum(outcome3[m:m + 24, 0])
                    if self.calibration_setting["NaturalGas_data"] == "Yes":
                        manipulated_result[k, 1] = np.sum(outcome3[m:m + 24, 1])
                    k += 1

            elif self.calibration_setting["Data_interval"] == "Hourly":
                if self.number_of_data == 2:
                    manipulated_result = np.zeros((8760, 2))
                else:
                    manipulated_result = np.zeros((8760, 1))

                if self.calibration_setting["Elec_data"] == "Yes":
                    manipulated_result[:, 0] = outcome3[:, 0]
                if self.calibration_setting["NaturalGas_data"] == "Yes":
                    manipulated_result[:, 1] = outcome3[:, 1]

            # 3. Caculate cvRMSE
            self.y_max = np.max(self.measuredData)
            if np.max(manipulated_result) > self.y_max:
                self.y_max = np.max(manipulated_result)
            deviation = 0
            for j in range(self.num_of_loop):
                if self.calibration_setting["Elec_data"] == "Yes":
                    deviation += (self.measuredData[j, 0] - manipulated_result[j, 0]) ** 2
                if self.calibration_setting["NaturalGas_data"] == "Yes":
                    deviation += (self.measuredData[j, 1] - manipulated_result[j, 1]) ** 2
            cvRMSE_compilation.append((1 / self.y_bar) * sqrt(deviation / (self.num_of_loop * self.number_of_data - 1)) * 100)
        #        print(f"cvRMSE: {cvRMSE_compilation}")
        #             diff = (self.measuredData[:,0] - manipulated_result[:,0])**2 / 11
        #             out = np.sqrt(np.sum(diff))/self.y_bar
        #             print('cvRMSE is:')
        #             print(out*100)

        return cvRMSE_compilation

    def CapX_Iteration(self, population):
        overall_budge_compilation = [];
        delivered_energy_compilation = []
        for i in range(self.genetic_algorithm_setting["num_of_population"]):
            self.JSON_Modification(population, i)

            j = 0;
            CapX_parameter_upfront_cost = 0;
            energy_cost = 0
            while j < self.num_of_parameters:
                if self.CapX_input_parameters[self.CapX_parameters[j]] == "Continuous":
                    # linear interpolation
                    # this has potential to be changed if and when we gain more cost data but in the meantime, it is a linear
                    # interpolation of the min and max cost we have
                    CapX_parameter_upfront_cost += self.Linear_Interpolation(
                        self.CapX_param_table_info[self.CapX_parameters[j]][0]["Param1"],
                        self.CapX_param_table_info[self.CapX_parameters[j]][1]["Param1"], \
                        self.CapX_param_table_info[self.CapX_parameters[j]][0]["Cost"],
                        self.CapX_param_table_info[self.CapX_parameters[j]][1]["Cost"], \
                        population[i, j])
                    j += 1
                elif self.CapX_input_parameters[self.CapX_parameters[j]] == "Discrete":
                    if self.CapX_parameters[j] == "Heating and Cooling Plants efficiencies (COPs)":
                        for element in self.CapX_param_table_info[self.CapX_parameters[j]]:
                            if element["Param1"] == population[i, j] and element["Param2"] == population[i, j + 1]:
                                CapX_parameter_upfront_cost += element["Cost"]
                                break
                        j += 2
                    elif self.CapX_parameters[j] in ["Roof1", "Opaque1", "Window1", "Natural Ventilation",
                                                     "Electric Battery"]:
                        for element in self.CapX_param_table_info[self.CapX_parameters[j]]:
                            if element["Param1"] == population[i, j] and element["Param2"] == population[i, j + 1] and \
                                    element["Param3"] == population[i, j + 2]:
                                CapX_parameter_upfront_cost += element["Cost"]
                                break
                        j += 3
                    else:
                        for element in self.CapX_param_table_info[self.CapX_parameters[j]]:
                            if element["Param1"] == population[i, j]:
                                CapX_parameter_upfront_cost += element["Cost"]
                                break
                        j += 1

            # 2. Re-calculations
            self.readJSON()
            self.GeneralDataProcessing()
            self.TransmissionHeatTransfer()
            self.VentilationAirFlowandVentilationHeatTransfer()
            self.PumpSystemEnergy()
            self.DHWandSolarWaterHeating()
            outcome, outcome2, outcome3 = self.hourly_BEM()  # outcome = self.deliveredEnergy, outcome2=self.Overall_deliveredEnergy, outcome3=self.deliveredEnergy_fuel

            delivered_energy_compilation.append(outcome2)

            # 3. Caculate Overall budget (= CapX initial upfront cost + energy cost*discount factor)
            energy_cost = np.sum(outcome3[:, 0]) * self.totalArea * self.CapX_setting["Elec_Cost"] + np.sum(outcome3[:, 1]) * self.totalArea * self.CapX_setting["NaturalGas_Cost"]

            overall_budge_compilation.append(energy_cost*self.interest_rate + CapX_parameter_upfront_cost)
            # overall_budge_compilation.append(CapX_parameter_upfront_cost)
        print(f"overall_budge_compilation: {overall_budge_compilation}")
        return overall_budge_compilation, delivered_energy_compilation

    def Crossover_Mutation(self, population):
        """
        1. Crossover: i) shuffle the mating pool's row
                     ii) cross over the first 2 rows in order iii) choose ONE croosing site randomly between 1 and L-1 according to a uniform distribution.
        2. Mutation: mutate the offspring of crossover directly
        """
        crossover1 = [];
        crossover2 = [];
        multiple_param_int = 0
        num_pop = population.shape[0]
        for i in range(0, int(num_pop) - 1, 2):
            if random() <= self.genetic_algorithm_setting["crossover_rate"] and self.num_of_parameters > 1:
                # Choose randomly the crossover_location
                crossover_location = randint(1, self.num_of_parameters - 1)
                crossover1 = list(population[i])
                crossover2 = list(population[i + 1])
                temporary = crossover1[crossover_location:self.num_of_parameters]

                crossover1[crossover_location:self.num_of_parameters] = crossover2[
                                                                        crossover_location:self.num_of_parameters]
                crossover2[crossover_location:self.num_of_parameters] = temporary

                population[i, :] = crossover1;
                population[i + 1, :] = crossover2
        ##                    print(f"crossover i: {i}, location: {crossover_location}")
        n_mutations = 0
        if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
            for j in range(self.num_of_parameters):
                for k in range(int(num_pop)):
                    if random() <= self.genetic_algorithm_setting["mutation_rate"]:
                        n_mutations += 1
                        if self.calibration_param_info[self.calibration_parameters[j]][
                            "Type"] == "Integer":  # For the schedule case, it's always float type, so it's fine to use this conditional statement
                            population[k, j] = randint(
                                self.calibration_param_info[self.calibration_parameters[j]]["Min"],
                                self.calibration_param_info[self.calibration_parameters[j]]["Max"])
                        else:
                            population[k, j] = uniform(
                                self.calibration_param_info[self.calibration_parameters[j]]["Min"],
                                self.calibration_param_info[self.calibration_parameters[j]]["Max"])

        else:  # CapX
            for k in range(int(num_pop)):
                j = 0;
                tempo_int = 0
                while j < self.num_of_parameters:
                    if random() <= self.genetic_algorithm_setting["mutation_rate"]:
                        n_mutations += 1
                        if self.CapX_input_parameters[self.CapX_parameters[j]] == "Continuous":
                            population[k, j] = uniform(self.CapX_param_table_info[self.CapX_parameters[j]][0]["Param1"],
                                                       self.CapX_param_table_info[self.CapX_parameters[j]][1]["Param1"])
                            j += 1
                        elif self.CapX_input_parameters[self.CapX_parameters[j]] == "Discrete":
                            tempo_int = randint(0, len(self.CapX_param_table_info[self.CapX_parameters[j]]) - 1)
                            if self.CapX_parameters[j] == "Heating and Cooling Plants efficiencies (COPs)":
                                population[k, j] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                    "Param1"]
                                population[k, j + 1] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                    "Param2"]
                                j += 2
                            elif self.CapX_parameters[j] in ["Roof1", "Opaque1", "Window1", "Natural Ventilation",
                                                             "Electric Battery"]:
                                population[k, j] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                    "Param1"]
                                population[k, j + 1] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                    "Param2"]
                                population[k, j + 2] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                    "Param3"]
                                j += 3
                            else:
                                population[k, j] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                    "Param1"]
                                j += 1
                    ##                        print(f"mutation  i:{i} j:{j}")
                    else:
                        if self.CapX_parameters[j] == "Heating and Cooling Plants efficiencies (COPs)":
                            j += 2
                        elif self.CapX_parameters[j] in ["Roof1", "Opaque1", "Window1", "Natural Ventilation",
                                                         "Electric Battery"]:
                            j += 3
                        else:
                            j += 1

        print(f'number of mutations is {n_mutations}')
        return population

    def Genetic_Algorithm_Loop(self):
        best_so_far_fitness_value = []  # Save each generation's best fitness value
        best_so_far_chromosome = []  # Save each generation's best chromosome gene
        RWS = []  # For selection, RWS = roulette wheel selection
        RSW_entire_row_numbers = [i for i in range(self.genetic_algorithm_setting["num_of_population"])]
        RSW_selected_row_numbers = []
        loop_count = 0;
        converted_loop_count = 1
        offspring = np.zeros((self.genetic_algorithm_setting["num_of_population"], self.num_of_parameters))

        if self.genetic_algorithm_setting["random_seed"] == "None":
            pass
        else:
            seed(self.genetic_algorithm_setting["random_seed"])
            np.random.seed(self.genetic_algorithm_setting["random_seed"])

        # Generate INITIAL population
        if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
            for i in range(self.genetic_algorithm_setting["num_of_population"]):
                for j in range(self.num_of_parameters):
                    if self.calibration_param_info[self.calibration_parameters[j]]["Type"] == "Integer":
                        offspring[i, j] = randint(self.calibration_param_info[self.calibration_parameters[j]]["Min"],
                                                  self.calibration_param_info[self.calibration_parameters[j]]["Max"])
                    else:  # FLOAT or temperature setpint or schedule
                        offspring[i, j] = uniform(self.calibration_param_info[self.calibration_parameters[j]]["Min"],
                                                  self.calibration_param_info[self.calibration_parameters[j]]["Max"])
            # with open(self.buildingName) as f:
            #     data = json.load(f)
            data = self.jsonData
            initial_chrom = []
            for j, name in enumerate(self.calibration_parameters):
                if name == "Heating COP":
                    initial_chrom.append(data["HeatingCOP"])
                elif name == "Cooling COP":
                    initial_chrom.append(data["CoolingCOP"])
                elif name == "Extra ventilation above fresh air supply":
                    initial_chrom.append(data["ExtraVentilation"])
                elif name == "Heat Recovery Type":
                    initial_chrom.append(data["HeatRecoveryType"])
                elif name == "Exhaust Air Recirculation Percentage":
                    initial_chrom.append(data["ExhaustAirRecirculation"])
                elif name == "Building air leakage level":
                    initial_chrom.append(data["AirLeakageLevel"])
                elif name == "Specific Fan Power":
                    initial_chrom.append(data["SpecificFanPower"])
                elif name == "Fan Control Factor":
                    initial_chrom.append(data["FanControlFactor"])
                elif name == "DHW Generation System":
                    initial_chrom.append(data["DHWGenerationSystem"])
                elif name == "Occupancy":
                    initial_chrom.append(data["Zone1"]["Occupancy"])
                elif name == "Appliance":
                    initial_chrom.append(data["Zone1"]["Appliance"])
                elif name == "Lighting":
                    initial_chrom.append(data["Zone1"]["Lighting"])
                elif name == "Outdoor Air":
                    initial_chrom.append(data["Zone1"]["OutdoorAir"])
                elif name == "DHW":
                    initial_chrom.append(data["Zone1"]["DHW"])

                elif name == "Roof1 Uvalue":
                    initial_chrom.append(data["Material"]["Roof1"]["UValue"])
                elif name == "Roof1 Absorption coefficient":
                    initial_chrom.append(data["Material"]["Roof1"]["Absorptivity"])
                elif name == "Roof1 Emissivity":
                    initial_chrom.append(data["Material"]["Roof1"]["Emissivity"])

                elif name == "Opaque1 Uvalue":
                    initial_chrom.append(data["Material"]["Opaque1"]["UValue"])
                elif name == "Opaque1 Absorption coefficient":
                    initial_chrom.append(data["Material"]["Opaque1"]["Absorptivity"])
                elif name == "Opaque1 Emissivity":
                    initial_chrom.append(data["Material"]["Opaque1"]["Emissivity"])

                elif name == "Window1 Uvalue":
                    initial_chrom.append(data["Material"]["Window1"]["UValue"])
                elif name == "Window1 Emissivity":
                    initial_chrom.append(data["Material"]["Window1"]["Emissivity"])
                elif name == "Window1 Solar transmittance":
                    initial_chrom.append(data["Material"]["Window1"]["SolarTrasmittance"])

                elif name[:11] == "Window1 SRF":
                    if name == "Window1 SRF All":
                        initial_chrom.append(data["Envelope"]["Window1"]["S"]["SRF"])
                    elif name == "Window1 SRF_S":
                        initial_chrom.append(data["Envelope"]["Window1"]["S"]["SRF"])
                    elif name == "Window1 SRF_SE":
                        initial_chrom.append(data["Envelope"]["Window1"]["SE"]["SRF"])
                    elif name == "Window1 SRF_E":
                        initial_chrom.append(data["Envelope"]["Window1"]["E"]["SRF"])
                    elif name == "Window1 SRF_NE":
                        initial_chrom.append(data["Envelope"]["Window1"]["NE"]["SRF"])
                    elif name == "Window1 SRF_N":
                        initial_chrom.append(data["Envelope"]["Window1"]["N"]["SRF"])
                    elif name == "Window1 SRF_NW":
                        initial_chrom.append(data["Envelope"]["Window1"]["NW"]["SRF"])
                    elif name == "Window1 SRF_W":
                        initial_chrom.append(data["Envelope"]["Window1"]["W"]["SRF"])
                    elif name == "Window1 SRF_SW":
                        initial_chrom.append(data["Envelope"]["Window1"]["SW"]["SRF"])

                elif name == "Retail Refri Capacity":
                    initial_chrom.append(data["RetailRefrig"]["Capacity"])

                elif name == "EV Charger Power":
                    initial_chrom.append(data["ElectricVehicle"]["ChargingPower"])

                elif name == "Garage Appliance":
                    initial_chrom.append(data["Garage"]["Appliance"])

                elif name == "Garage Lighting":
                    initial_chrom.append(data["Garage"]["Lighting"])

                elif name == "Lighting Dimmer Weekday":
                    initial_chrom.append(data["LightingDimmer"]["DimmerSaving_WD"])

                elif name == "Lighting Dimmer Weekend":
                    initial_chrom.append(data["LightingDimmer"]["DimmerSaving_WE"])

                elif "Monthly" in name:  ####
                    if name[:9] == "Occupancy":
                        monthly_type = "Occupancy"
                    elif name[:9] == "Appliance":
                        monthly_type = "Appliance"
                    elif name[:8] == "Lighting":
                        monthly_type = "Lighting"

                    if self.month_list.index(self.calibration_param_info[name]["Month_Min"]) == self.month_list.index(
                            self.calibration_param_info[name]["Month_Max"]):
                        month = self.calibration_param_info[name]["Month_Min"]
                        initial_chrom.append(data["Monthly_InternalHeatGain"][monthly_type][month])

                    elif self.month_list.index(self.calibration_param_info[name]["Month_Min"]) < self.month_list.index(
                            self.calibration_param_info[name]["Month_Max"]):
                        # print(f"calibration_param_info[name]: {self.calibration_param_info[name]}")
                        z = self.month_list.index(self.calibration_param_info[name]["Month_Min"]) + 1
                        initial_chrom.append(data["Monthly_InternalHeatGain"][monthly_type][self.month_list[z]])
                    else:
                        raise Exception(
                            "Please check the 'Monthly_Min' and 'Monthly_Max' in the 'Monthly Internal Heat Gains' table.")
                #TODO: Look into set point
                elif isinstance(int(name), int)== True: # Temperature setpoint or Schedule
                    hour_max = self.calibration_param_info[name]["Hour_Max"]
                    hour_min = self.calibration_param_info[name]["Hour_Min"]
                    if self.calibration_param_info[name]["Hour_Max"] < self.calibration_param_info[name][
                        "Hour_Min"]:
                        hour_list = list(range(hour_min, 25)) + list(range(1, hour_max + 1))
                        hour_list.sort()
                    else:
                        hour_list = list(range(hour_min, hour_max + 1))
                    if self.calibration_param_info[name]["Name"] in ["Heating Temperature Setpoint", "Cooilng Temperature Setpoint"]:
                        k = hour_list[0]
                        temp = data["TemperatureSetPoint"]["".join(["Hour",str(k)])]["".join([self.calibration_param_info[name]["Name"][:7], "_", self.calibration_param_info[name]["Type"]])]## added '_'
                        initial_chrom.append(temp)
                    else: # Occupancy, appliance, lighting
                        k = hour_list[0]
                        temp = data["Zone1_Schedule"]["".join(["Hour",str(k)])]["".join([self.calibration_param_info[name]["Name"][:-9], "_", self.calibration_param_info[name]["Type"]])]
                        initial_chrom.append(temp)
            offspring[0] = initial_chrom

        else:  # CapX
            for i in range(self.genetic_algorithm_setting["num_of_population"]):
                j = 0
                while j < self.num_of_parameters:
                    if self.CapX_input_parameters[self.CapX_parameters[j]] == "Continuous":
                        offspring[i, j] = uniform(self.CapX_param_table_info[self.CapX_parameters[j]][0]["Param1"],
                                                  self.CapX_param_table_info[self.CapX_parameters[j]][1]["Param1"])
                        j += 1
                    elif self.CapX_input_parameters[self.CapX_parameters[j]] == "Discrete":
                        tempo_int = randint(0, len(self.CapX_param_table_info[self.CapX_parameters[j]]) - 1)
                        if self.CapX_parameters[j] == "Heating and Cooling Plants efficiencies (COPs)":
                            offspring[i, j] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int]["Param1"]
                            offspring[i, j + 1] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                "Param2"]
                            j += 2
                        elif self.CapX_parameters[j] in ["Roof1", "Opaque1", "Window1", "Natural Ventilation",
                                                         "Electric Battery"]:
                            offspring[i, j] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int]["Param1"]
                            offspring[i, j + 1] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                "Param2"]
                            offspring[i, j + 2] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int][
                                "Param3"]
                            j += 3
                        else:
                            offspring[i, j] = self.CapX_param_table_info[self.CapX_parameters[j]][tempo_int]["Param1"]
                            j += 1

        # Start the calibration loop
        start = timer()
        old_lowest = []
        old_cv = []
        while 1:
            if loop_count > 0:
                offspring = selected_offspring
            # print(f"offspring:\n{offspring}")
            print('------------------------------------------')

            if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
                # Call the 'Calibration_Iteration' to evaluate each chromosome
                t0 = time.time()
                evaluation_from_iteration_method = self.Calibration_Iteration(offspring)
                t1 = time.time()
                # print(t1-t0)

            else:
                evaluation_from_iteration_method, delivered_energy_comp = self.CapX_Iteration(offspring)

            print(f"evaluation_from_iteration_method:\n{evaluation_from_iteration_method}")

            min_evaluation_from_iteration_method = min(evaluation_from_iteration_method)
            sum_evaluation_from_iteration_method = sum(evaluation_from_iteration_method)

            if self.genetic_algorithm_setting["calibration_or_CapX"] == "CapX":
                min_delivered_energy = delivered_energy_comp[
                    evaluation_from_iteration_method.index(min_evaluation_from_iteration_method)]

            # Save the minimum fitness value
            best_so_far_fitness_value.append(min_evaluation_from_iteration_method)
            print(f"\nbest fitnessvalue at each generation:\n{best_so_far_fitness_value}")

            # Save the minimum fitness value's chromosome
            best_chromosome_index = evaluation_from_iteration_method.index(min_evaluation_from_iteration_method)
            best_so_far_chromosome.append(list(offspring[best_chromosome_index, :]))
            print(f"best_so_far_chromosome:\n{best_so_far_chromosome}")

            # Make a decision on the looping
            loop_count += 1

            if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
                if abs(best_so_far_fitness_value[loop_count - 1] - best_so_far_fitness_value[loop_count - 2]) > \
                        self.calibration_setting["Convergence"]:
                    converted_loop_count = 1
                    print(f'CVRMSE has improved! and it is now {round(best_so_far_fitness_value[-1],3)}%')
                if loop_count > 1:
                    if min_evaluation_from_iteration_method < self.calibration_setting["cvRMSE_criterion"]:
                        print(
                            f"Calibration is completed - The input cvRMSE criterion {self.calibration_setting['cvRMSE_criterion']}% is met.")
                        break
                    elif abs(best_so_far_fitness_value[loop_count - 1] - best_so_far_fitness_value[loop_count - 2]) <= \
                            self.calibration_setting["Convergence"]:

                        converted_loop_count += 1
                        print(f"CV RMSE of {round(best_so_far_fitness_value[-1],3)}% has repeated for the {converted_loop_count}th time")
                        if converted_loop_count == self.calibration_setting["N_times_Convergence"]:
                            print(
                                f"Calibration is completed - The cvRMSE value remained constant within {self.calibration_setting['Convergence']} for {self.calibration_setting['N_times_Convergence']} times.")
                            break
                        elif converted_loop_count > np.round(self.calibration_setting["N_times_Convergence"] * 2/3):
                            self.genetic_algorithm_setting["mutation_rate"] = .6
                            print(f"the new mutation rate is {self.genetic_algorithm_setting['mutation_rate']}")
            else:  # CapX
                if min_evaluation_from_iteration_method <= self.CapX_setting["Budget"] and min_delivered_energy <= \
                        self.CapX_setting["Delivered_Energy"] and loop_count > 1:
                    print(
                        f"CapX is completed - The input budget ${self.CapX_setting['Budget']} and min delivered energy kWh/m2/hr {self.CapX_setting['Delivered_Energy']} constraints are met.")
                    break

            end = timer()
            if (end - start) / 60 > self.genetic_algorithm_setting["Code_Execution_Time"]:
                print(
                    f"optimization is completed - Code elapsed time reached the input limitation time {self.genetic_algorithm_setting['Code_Execution_Time']} mins.")
                break

            ##            print(f"\noffspring1:\n{offspring}")
            top_per_selection = self.genetic_algorithm_setting["top_percentage_selection"] / 100
            # print(top_per_selection)
            lowest = int(np.ceil(self.genetic_algorithm_setting["num_of_population"] * top_per_selection))
            # Pick the N lowest cvRMSE value indices

            lowest_row_numbers = np.argsort(evaluation_from_iteration_method)[:lowest]
            # if len(old_lowest) != 0:
            #     print("The new top chromosomes are:")
            #     print((lowest_row_numbers))
            #     print("with the following fitness:")
            #     print(np.sort(evaluation_from_iteration_method))
            #     print("The old top chromosomes are:")
            #     print((old_lowest))
            #     print("with the following fitness:")
            #     print(np.sort(old_cv))

            old_lowest = lowest_row_numbers
            old_cv = evaluation_from_iteration_method

            ##            print(f"\nlowest_row_numbers:  {lowest_row_numbers}")

            # Selection (Roulette Wheel Selection (RWS))
            reverse_cvRMSE_sum = sum(
                [sum_evaluation_from_iteration_method / x for x in evaluation_from_iteration_method])
            RWS = [(sum_evaluation_from_iteration_method / x) / reverse_cvRMSE_sum for x in
                   evaluation_from_iteration_method]

            RSW_selected_row_numbers = np.random.choice(RSW_entire_row_numbers,
                                                        self.genetic_algorithm_setting["num_of_population"] - lowest,
                                                        replace=True, p=RWS)
            ##            print(f"\nRWS: {RWS}")
            ##            print(f"RSW_selected_row_numbers: {RSW_selected_row_numbers}")

            # Copy the fist 2 the population with the best chromosomes from the previous generation to prevent from increasing the fitness value
            selected_offspring = np.zeros((self.genetic_algorithm_setting["num_of_population"], self.num_of_parameters))
            for i, k in enumerate(lowest_row_numbers):
                selected_offspring[i, :] = offspring[k, :]

            # Copy the population with the best chromosomes from the current generation
            k = int(np.ceil(self.genetic_algorithm_setting["num_of_population"] * top_per_selection))
            half = k
            for i in RSW_selected_row_numbers:
                selected_offspring[k, :] = offspring[i, :]
                k += 1
            ##            print(f"\nselected_offspring2:\n{selected_offspring}")

            # Shuffle the rows to prepare for the crossover
            np.random.shuffle(selected_offspring[half:])
            ##            print(f"\nselected_offspring3:\n{selected_offspring}")

            # Crossover Mutation
            selected_offspring[half:] = self.Crossover_Mutation(selected_offspring[half:])
        ##            print(f"\nselected_offspring4:\n{selected_offspring}")

        # Record the calibrated values in the "Calibration_Result.csv" file
        with open(f"./Output/Genetic_Algorithm_Result_{self.result_file_name}.csv", 'w', newline='') as csvfile:
            # creating a csv writer object
            csvwriter = csv.writer(csvfile)

            # writing the fields
            if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
                csvwriter.writerow(["cvRMSE Value"])
            else:
                csvwriter.writerow(["Net Present Cost [$]"])
            csvwriter.writerow(best_so_far_fitness_value)

            csvwriter.writerow([])
            csvwriter.writerow(["best chromosome in each generation"])
            if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
                csvwriter.writerow(self.calibration_parameters)
            else:
                csvwriter.writerow(self.CapX_parameters)

            for k in best_so_far_chromosome:
                csvwriter.writerow(k)
        # def visualize_best(self):
        # file = openpyxl.load_workbook('./Input/BEM_Optimization_Input_' + original_file_name + '.xlsx', data_only=True)
        # file_sheet = file['GeneticAlgorithm_Setting']
        # self.genetic_algorithm_setting["calibration_or_CapX"] = file_sheet.cell(row=7, column=3).value
        if self.genetic_algorithm_setting["calibration_or_CapX"] == "Calibration":
            best = np.zeros((2, self.num_of_parameters))
            best[0, :] = best_so_far_chromosome[-1]

            # building_name = self.buildingName
            # self.JSON_Modification(best, 0, building_name, calibrated = True)
            self.JSON_Modification(best, 0, True)
            # 2. Re-calculations (Assumed BEM calculation was conducted at least once before the "cvRMSE" was executed.)
            # self.readJSON(building_name)
            self.readJSON()

            self.GeneralDataProcessing()
            self.TransmissionHeatTransfer()
            self.VentilationAirFlowandVentilationHeatTransfer()
            self.PumpSystemEnergy()
            self.DHWandSolarWaterHeating()

            outcome, outcome2, outcome3, grouped = self.hourly_BEM()
            out = np.asarray(outcome[:, -1]) * self.totalArea / 1000

            manipulated_result = np.zeros((12, 1))
            manipulated_result[0, 0] = np.sum(outcome3[0:744, 0])
            manipulated_result[1, 0] = np.sum(outcome3[744:1416, 0])
            manipulated_result[2, 0] = np.sum(outcome3[1416:2159, 0])
            manipulated_result[3, 0] = np.sum(outcome3[2159:2880, 0])
            manipulated_result[4, 0] = np.sum(outcome3[2880:3624, 0])
            manipulated_result[5, 0] = np.sum(outcome3[3624:4344, 0])
            manipulated_result[6, 0] = np.sum(outcome3[4344:5088, 0])
            manipulated_result[7, 0] = np.sum(outcome3[5088:5832, 0])
            manipulated_result[8, 0] = np.sum(outcome3[5832:6552, 0])
            manipulated_result[9, 0] = np.sum(outcome3[6552:7296, 0])
            manipulated_result[10, 0] = np.sum(outcome3[7296:8016, 0])
            manipulated_result[11, 0] = np.sum(outcome3[8016:8760, 0])

            # deviation = 0
            # for j in range(self.num_of_loop):
            #     deviation += (self.measuredData[j, 0] - manipulated_result[j, 0]) ** 2
            # cvRMSE = (1 / self.y_bar) * sqrt(deviation / (self.num_of_loop * self.number_of_data - 1)) * 100

            data = pd.DataFrame(out, columns=['Delivered'])

            # Outputting Simulated Data into Monthly, Daily, Hourly Format

            if self.calibration_setting["Data_interval"] == "Monthly":
                data['Month'] = 'blank'
                data.loc[:743, 'Month'] = 'January'
                data.loc[744:1415, 'Month'] = 'February'
                data.loc[1416:2159, 'Month'] = 'March'
                data.loc[2160:2879, 'Month'] = 'April'
                data.loc[2880:3623, 'Month'] = 'May'
                data.loc[3624:4343, 'Month'] = 'June'
                data.loc[4344:5087, 'Month'] = 'July'
                data.loc[5088:5831, 'Month'] = 'August'
                data.loc[5832:6551, 'Month'] = 'Septemeber'
                data.loc[6552:7295, 'Month'] = 'October'
                data.loc[7296:8015, 'Month'] = 'November'
                data.loc[8016:8759, 'Month'] = 'December'
                grouped = (data.groupby(['Month'], sort=False).sum()).reset_index()
                self.simulated = grouped.Delivered.values.tolist()
            elif self.calibration_setting["Data_interval"] == "Hourly":
                self.delivered = pd.DataFrame(out, columns=['Delivered'])
                self.simulated = self.delivered.Delivered.values.tolist()
            # plt.figure(figsize=(14, 8))
            # plt.plot(grouped.Month.values, grouped.Delivered.values, label="Model", marker='o')
            # plt.plot(grouped.Month.values, self.measuredData[:, 0], label="Utility", marker='o')
            # plt.ylim(0, (self.y_max + 100000))
            # plt.legend()
            # plt.xlabel('Month')
            # plt.ylabel('Energy Consumption (kWh)')
            # plt.title(f"Model Results vs. Utility Data for CV RMSE of {round(best_so_far_fitness_value[-1],3)}%")
            # plt.ticklabel_format(style='plain', axis='y')
            # plt.savefig(f"calibration_result_{self.result_file_name}.png")
            # plt.show()

            return self.simulated, list(self.measuredData[:, 0]), self.calibration_setting["Data_interval"]


        else:
            best = np.zeros((2, self.num_of_parameters))
            best[0, :] = best_so_far_chromosome[-1]

            # building_name = self.buildingName
            # self.JSON_Modification(best, 0, building_name, calibrated = True)
            self.JSON_Modification(best, 0, True)
            # 2. Re-calculations (Assumed BEM calculation was conducted at least once before the "cvRMSE" was executed.)
            # self.readJSON(building_name)
            self.readJSON()

            self.GeneralDataProcessing()
            self.TransmissionHeatTransfer()
            self.VentilationAirFlowandVentilationHeatTransfer()
            self.PumpSystemEnergy()
            self.DHWandSolarWaterHeating()

            outcome, outcome2, outcome3, grouped = self.hourly_BEM()
            out = np.asarray(outcome[:, -1]) * self.totalArea / 1000

            manipulated_result = np.zeros((12, 1))
            manipulated_result[0, 0] = np.sum(outcome3[0:744, 0])
            manipulated_result[1, 0] = np.sum(outcome3[744:1416, 0])
            manipulated_result[2, 0] = np.sum(outcome3[1416:2159, 0])
            manipulated_result[3, 0] = np.sum(outcome3[2159:2880, 0])
            manipulated_result[4, 0] = np.sum(outcome3[2880:3624, 0])
            manipulated_result[5, 0] = np.sum(outcome3[3624:4344, 0])
            manipulated_result[6, 0] = np.sum(outcome3[4344:5088, 0])
            manipulated_result[7, 0] = np.sum(outcome3[5088:5832, 0])
            manipulated_result[8, 0] = np.sum(outcome3[5832:6552, 0])
            manipulated_result[9, 0] = np.sum(outcome3[6552:7296, 0])
            manipulated_result[10, 0] = np.sum(outcome3[7296:8016, 0])
            manipulated_result[11, 0] = np.sum(outcome3[8016:8760, 0])

            data = pd.DataFrame(out, columns=['Delivered'])

            if self.calibration_setting["Data_interval"] == "Monthly":
                data['Month'] = 'blank'
                data.loc[:743, 'Month'] = 'January'
                data.loc[744:1415, 'Month'] = 'February'
                data.loc[1416:2159, 'Month'] = 'March'
                data.loc[2160:2879, 'Month'] = 'April'
                data.loc[2880:3623, 'Month'] = 'May'
                data.loc[3624:4343, 'Month'] = 'June'
                data.loc[4344:5087, 'Month'] = 'July'
                data.loc[5088:5831, 'Month'] = 'August'
                data.loc[5832:6551, 'Month'] = 'Septemeber'
                data.loc[6552:7295, 'Month'] = 'October'
                data.loc[7296:8015, 'Month'] = 'November'
                data.loc[8016:8759, 'Month'] = 'December'
                grouped = (data.groupby(['Month'], sort=False).sum()).reset_index()
                self.simulated = grouped.Delivered.values.tolist()
            elif self.calibration_setting["Data_interval"] == "Hourly":
                self.delivered = pd.DataFrame(out, columns=['Delivered'])
                self.simulated = self.delivered.Delivered.values.tolist()
            # data['Month'] = 'blank'
            # data.loc[:743, 'Month'] = 'January'
            # data.loc[744:1415, 'Month'] = 'February'
            # data.loc[1416:2159, 'Month'] = 'March'
            # data.loc[2160:2879, 'Month'] = 'April'
            # data.loc[2880:3623, 'Month'] = 'May'
            # data.loc[3624:4343, 'Month'] = 'June'
            # data.loc[4344:5087, 'Month'] = 'July'
            # data.loc[5088:5831, 'Month'] = 'August'
            # data.loc[5832:6551, 'Month'] = 'Septemeber'
            # data.loc[6552:7295, 'Month'] = 'October'
            # data.loc[7296:8015, 'Month'] = 'November'
            # data.loc[8016:8759, 'Month'] = 'December'
            # grouped = (data.groupby(['Month'], sort=False).sum()).reset_index()
            # plt.figure(figsize=(14, 8))
            # plt.plot(grouped.Month.values, grouped.Delivered.values, label="Model", marker='o')

            # plt.legend()
            # plt.xlabel('Month')
            # plt.ylabel('Energy Consumption (kWh)')
            # plt.title(f"Model Results vs. Utility Data for CV RMSE of {round(best_so_far_fitness_value[-1],3)}%")
            # plt.ticklabel_format(style='plain', axis='y')
            # plt.savefig(f"calibration_result_{self.result_file_name}.png")
            # plt.show()

            self.grouped = grouped.Delivered.values

            return self.simulated, list(self.measuredData[:, 0]), self.calibration_setting["Data_interval"]


        # for j in range(self.num_of_loop):
        #     if self.calibration_setting["Elec_data"] == "Yes":
        #         deviation += (self.measuredData[j,0]-manipulated_result[j,0])**2
        #     if self.calibration_setting["NaturalGas_data"] == "Yes":
        #         deviation += (self.measuredData[j,1]-manipulated_result[j,1])**2


def BEMP_Optimization(buildingName, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Esol_30, Esol_45, Esol_60, Esol_90,
                      original_file_name, result_file_name):
    instance = BEMP_Calibration_CapX(buildingName, weatherData, SRF_overhang, SRF_fin, SRF_horizon, Esol_30, Esol_45,
                                     Esol_60, Esol_90, original_file_name, result_file_name)
    instance.readJSON()
    instance.GeneralDataProcessing()
    instance.TransmissionHeatTransfer()
    instance.VentilationAirFlowandVentilationHeatTransfer()
    instance.BEMSystem()
    instance.CoolingandHeatingSystemEnergy()
    instance.FanEnergy()
    instance.PumpSystemEnergy()
    instance.DHWandSolarWaterHeating()
    outcome, outcome2, outcome3, grouped = instance.hourly_BEM()
    real, simulated, interval = instance.Genetic_Algorithm_Loop()

    return simulated, real, interval


if __name__ == '__main__':
    print("Please execute the 'main.py' script.")
